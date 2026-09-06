# pages/5_Yillik_Menu.py
#
# Yillik Menu Uretim Motoru (ilk surum): global tarif kutuphanesinden
# (receteler, isletme_id NULL) anayasa kurallarina uygun ornek haftalik
# menu uretir. Henuz eklenmeyenler: kisisel_beslenme_profili filtrelemesi,
# menu_takvimi/menu_takvimi_ogeleri'ne yazma (sadece ekranda gosteriyor).

import datetime
import io
import json
import random

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# NOT (12 Agustos 2026, Oturum 11): logo artik burada AYRICA gosterilmiyor -- app.py'deki ozel menu satirinin icine tasindi, orada zaten her sayfa gecisinde render ediliyor. Burada tekrar cagirmak cift logoya yol acardi.

from db import get_supabase, oturumu_uygula
from besin_sabitleri import TUM_BESIN_ALANLARI, BESIN_ETIKET, BESIN_ARALIK, kanonik_sirala
from uretim_algoritmasi import hafta_olustur, ogun_olustur, _taban_kelime, _hedef_mesafesi, _fast_food_sec

# DOKSAN ALTINCI DUZELTME (4 Eylul 2026): TEMEL_5 dosyanin EN BASINA
# tasindi -- daha once _gun_popup_govdesini_ciz'in icinde (cok asagida)
# tanimliydi, TUM kullanim yerlerinden ONCE gelmesi gerekiyordu.
TEMEL_5 = {"kalori", "protein", "yag", "karbonhidrat", "gi"}

# DOKSAN ALTINCI DUZELTME DENEMESI (4 Eylul 2026, GERI ALINDI): Bahri,
# Kasım ayı üretiminde 84 öğünün 74'ünün (%88!) hedef dışı çıktığını
# bildirdi. Ilk teorim: uretim SADECE TEMEL_5'i hedeflemeli, genisletilmis
# 22 oge sadece haftalik ortalamayla (uretimden SONRA) kontrol edilmeli
# -- boylece uretim "gercekci olmayan" 27-eszamanli hedefle
# ugrasmayi biraksin. BU TEORI 10 farkli tohumla test edildi VE YANLIS
# CIKTI: TEMEL_5-only uretim, TAM hedefli uretimden DAHA KOTU sonuc
# verdi (%85.7 hedef disi, TAM hedefli %50.7'ye karsi) -- yani
# genisletilmis ogeleri uretim sirasinda TAMAMEN gormezden gelmek,
# onlari (kismen de olsa) hedeflemekten DAHA KOTU. Degisiklik bu
# yuzden GERI ALINDI, uretim hala TAM hedefi kullaniyor. Asil %88
# sorunu HALA COZULMEDI -- gercek sebep muhtemelen 27 eszamanli
# hedefin GERCEKTEN cok zor olmasi + uretim algoritmasinin bunu
# HAFTALIK bazda degil hala TEK OGUN bazinda optimize etmesi -- bu,
# tek-ogun uretiminin dogasi geregi COZMESI ZOR bir sorun, ayri ve
# daha derin bir tasarim calismasi gerektiriyor.

MEVSIM_AYLARI = {
    "kis": ["Aralık", "Ocak", "Şubat"],
    "ilkbahar": ["Mart", "Nisan", "Mayıs"],
    "yaz": ["Haziran", "Temmuz", "Ağustos"],
    "sonbahar": ["Eylül", "Ekim", "Kasım"],
}
AYLAR_SIRALI = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

# YIRMINCI DUZELTME (13 Agustos 2026, Oturum 11): kullanicinin "gercek
# takvim tarihi" istegi uzerine -- Ay -> Mevsim TERS eslesmesi (asagidaki
# gercek-hafta hesaplamasinda her GUNUN kendi mevsimini bulmak icin).
AY_TO_MEVSIM = {ay: mevsim for mevsim, aylar in MEVSIM_AYLARI.items() for ay in aylar}


def _ay_gercek_haftalari(yil, ay_adi):
    """Verilen (yil, ay) icin GERCEK ISO takvim haftalarini (Pazartesi-
    Pazar) dondurur -- ayin kendisi bir haftanin ORTASINDA baslayip
    bitebildigi icin, DONEN İLK ve SON hafta komsu aya tasan gunler
    icerebilir (ör. Agustos 2026'nin ilk haftasi 27 Temmuz Pazartesi'nde
    baslar, 2 Agustos Pazar'da biter -- kullanicinin gonderdigi gercek
    takvim gorseliyle birebir dogrulandi). Her hafta TAM 7 gercek
    datetime.date objesi icerir (Pazartesi=index 0 .. Pazar=index 6).
    Donen deger: [[date, date, ..., 7 tane], ...] seklinde bir hafta listesi."""
    ay_no = AYLAR_SIRALI.index(ay_adi) + 1
    ilk_gun = datetime.date(yil, ay_no, 1)
    if ay_no == 12:
        son_gun = datetime.date(yil, 12, 31)
    else:
        son_gun = datetime.date(yil, ay_no + 1, 1) - datetime.timedelta(days=1)

    # Ayin ilk gununun ait oldugu haftanin Pazartesi'si (weekday(): Pzt=0)
    ilk_hafta_pazartesi = ilk_gun - datetime.timedelta(days=ilk_gun.weekday())
    # Ayin son gununun ait oldugu haftanin Pazartesi'si
    son_hafta_pazartesi = son_gun - datetime.timedelta(days=son_gun.weekday())

    haftalar = []
    pazartesi = ilk_hafta_pazartesi
    while pazartesi <= son_hafta_pazartesi:
        hafta = [pazartesi + datetime.timedelta(days=i) for i in range(7)]
        haftalar.append(hafta)
        pazartesi += datetime.timedelta(weeks=1)
    return haftalar


def _tarih_mevsimi(tarih):
    """Bir datetime.date'in ait oldugu Turkce ay adindan mevsimini bulur."""
    ay_adi = AYLAR_SIRALI[tarih.month - 1]
    return AY_TO_MEVSIM[ay_adi]

# ON DOKUZUNCU DUZELTME (13 Agustos 2026, Oturum 11): malzemeler tablosu
# artik 27 ek besin ogesi iceriyor (45_genisletilmis_besin_degerleri.sql).
# Kullanicinin istegi: bunlarin da kalori ornegindeki gibi SECILEBILIR ve
# hedef araligiyla KISITLANABILIR olmasi -- ama hepsini otomatik gostermek
# DEGIL (31 satir birden cok kalabalik olurdu). Bu yuzden asagida TUM
# alanlarin (mevcut 5 + yeni 27) tek bir katalogu var; kullanici bunlardan
# HANGILERINI hedeflemek istedigini bir multiselect ile seciyor, sadece
# secilenler icin min/max girisi gosteriliyor (bkz. asagidaki
# "besin_hedefi_kullan" bolumu).
# Aralik/varsayilan degerleri kabaca "gunluk RDA'nin uc oguune bolunmus
# hali" mantigiyla, TEK BIR OGUN icin makul bir baslangic noktasi olarak
# secildi -- kesin bilimsel bir tavsiye degil, kullanici kendi mutfagina
# gore ayarlayabilir.
# malzemeler tablosundaki kolon adlari (5 temel alan disindakiler icin
# ayni isim, kolon adiyla anahtar birebir ayni secildi)
_GENISLETILMIS_KOLONLAR = [a for a, *_ in TUM_BESIN_ALANLARI if a not in ("kalori", "protein", "yag", "karbonhidrat", "gi")]

st.set_page_config(page_title="Aylık Menü", page_icon="assets/favicon.png", layout="wide")

supabase = get_supabase()
oturumu_uygula(supabase)

st.title("Aylık Menü Üretim Motoru")
st.caption(
    "Türk mutfağı tarif kütüphanesinden, anayasa kurallarına uygun "
    "(madde 8, 11, 13) örnek haftalık menü üretir. İlk sürüm — kişisel "
    "beslenme profili filtrelemesi ve takvime kaydetme henüz eklenmedi."
)


def _duyarli_sutun_css_uygula():
    """YIRMI SEKIZINCI DUZELTME (13 Agustos 2026, Oturum 11): asil kart
    stili (_yillik_menu_tasarim_stilini_uygula) SADECE menu uretildikten
    sonra (hafta kartlari cizilirken) cagriliyordu -- ama bolge menu
    kartlari (asagida) menu uretilmeden ONCE, sayfanin en basinda
    render ediliyor. Kullanici mobil/tablette bu erken kartlarin da
    (hafta kartlari gibi) dar ekranda sikisip metnin kelime ortasindan
    bolundugunu bildirdi. Bu kucuk, ERKEN cagrilan fonksiyon SADECE
    duyarli (responsive) sutun kuralini once yukluyor -- boylece
    hicbir kart turu (menu uretilmis olsun olmasin) bu korumadan
    mahrum kalmiyor. Asil buyuk stil fonksiyonu da ayni kurali tekrar
    icerir (zararsiz, CSS yeniden-enjeksiyonu idempotenttir)."""
    st.markdown(
        '<style>@media (max-width: 1024px) { [data-testid="stHorizontalBlock"] '
        '{ flex-wrap: wrap !important; row-gap: 10px; } '
        '[data-testid="stHorizontalBlock"] > div { flex: 1 1 30% !important; '
        'min-width: 130px !important; } }</style>',
        unsafe_allow_html=True,
    )


_duyarli_sutun_css_uygula()


@st.cache_data(ttl=3600)
def _mutfaklari_getir():
    return (supabase.table("mutfaklar").select("kod, ad").execute()).data


def _sayfalayarak_getir(sorgu_uret, sayfa_boyutu=1000):
    """OTUZUNCU DUZELTME (13 Agustos 2026, Oturum 11): kullanicinin
    bildirdigi maliyet tutarsizligi ARASTIRILDI -- kok sebep bulundu.
    Supabase/PostgREST, .range() belirtilmese bile sorgu basina
    varsayilan olarak en fazla 1000 satir donduruyor -- sinirin
    uzerindeki satirlar HATA VERMEDEN sessizce kesiliyor. Bu sayfanin
    fiyat sorgulari (asagida) bu korumayi hic kullanmiyordu --
    pages/5_Tarif_Kutuphanesi.py'de AYNI sorun icin zaten var olan bu
    yardimci fonksiyon buraya da kopyalandi ve HER IKI fiyat
    sorgusuna (global + isletmeye ozel tarifler) uygulandi. `sorgu_uret`,
    her cagrildiginda henuz .range()/.execute() uygulanmamis YENI bir
    sorgu builder'i donduren bir fonksiyon olmali (ayni builder tekrar
    kullanilamiyor)."""
    tumu = []
    offset = 0
    while True:
        sayfa = sorgu_uret().range(offset, offset + sayfa_boyutu - 1).execute().data
        tumu.extend(sayfa)
        if len(sayfa) < sayfa_boyutu:
            break
        offset += sayfa_boyutu
    return tumu


# OTUZ BIRINCI DUZELTME (13 Agustos 2026, Oturum 11): kullanicinin
# talebiyle -- pop-up'taki maliyet SADECE malzeme maliyetiydi, oysa
# Tarif Kutuphanesi'nde ("Gerçek üretim maliyeti") zaten var olan
# malzeme+enerji+işçilik modeliyle TUTARSIZDI. Asagidaki 3 fonksiyon,
# pages/5_Tarif_Kutuphanesi.py'deki BIREBIR AYNI mantikla buraya da
# eklendi, boylece pop-up da tam maliyeti gosterebiliyor.
@st.cache_data(ttl=3600)
def _tum_tarif_id_by_ad_getir(isletme_id):
    """Pop-up'ta uretim asamalarina (enerji/iscilik icin) erismek icin
    tarif ADINDAN id'sine ihtiyacimiz var -- detay sozlugu (yukarida)
    sadece ada gore anahtarlanmis, id tasimiyor."""
    genel = _sayfalayarak_getir(
        lambda: supabase.table("receteler").select("id, ad").is_("isletme_id", "null")
    )
    ozel = _sayfalayarak_getir(
        lambda: supabase.table("receteler").select("id, ad").eq("isletme_id", isletme_id)
    )
    return {r["ad"]: r["id"] for r in genel + ozel}


@st.cache_data(ttl=3600)
def _uretim_asamalarini_getir(recete_id):
    asamalar = (
        supabase.table("recete_asamalari")
        .select("id, ad, sira, sure_dakika, aktif_dakika, isil_islem_mi, enerji_kaynagi, baslangic_sicaklik, hedef_sicaklik, verimlilik_orani")
        .eq("recete_id", recete_id)
        .order("sira")
        .execute()
    ).data
    if not asamalar:
        return []
    asama_malzeme_kayitlari = (
        supabase.table("asama_malzemeleri")
        .select("asama_id, recete_malzemeleri(miktar_gram, malzemeler(ozgul_isi))")
        .in_("asama_id", [a["id"] for a in asamalar])
        .execute()
    ).data
    for a in asamalar:
        a["isitilan_kutle_gram"] = sum(
            (k["recete_malzemeleri"] or {}).get("miktar_gram", 0)
            for k in asama_malzeme_kayitlari
            if k["asama_id"] == a["id"]
        )
        a["agirlikli_ozgul_isi"] = None
        kayitlar_bu_asama = [
            k for k in asama_malzeme_kayitlari
            if k["asama_id"] == a["id"] and k.get("recete_malzemeleri")
        ]
        if kayitlar_bu_asama and a["isitilan_kutle_gram"] > 0:
            toplam_ozgul_isi_agirlikli = sum(
                k["recete_malzemeleri"]["miktar_gram"]
                * ((k["recete_malzemeleri"].get("malzemeler") or {}).get("ozgul_isi") or 0)
                for k in kayitlar_bu_asama
            )
            a["agirlikli_ozgul_isi"] = toplam_ozgul_isi_agirlikli / a["isitilan_kutle_gram"]
    return asamalar


@st.cache_data(ttl=3600)
def _maliyet_ayarlarini_getir(isletme_id):
    sonuc = (
        supabase.table("isletme_maliyet_ayarlari")
        .select("*")
        .eq("isletme_id", isletme_id)
        .execute()
    ).data
    if sonuc:
        return sonuc[0]
    return {
        "elektrik_birim_fiyat_eur_kwh": 0.12,
        "dogalgaz_birim_fiyat_eur_kwh": 0.08,
        "personel_saat_ucreti_eur": 5.0,
        "genel_gider_yuzdesi": 15.0,
    }


def _gercek_maliyet_hesapla(asamalar, ayarlar, porsiyon):
    enerji_eur = 0.0
    iscilik_dk = 0.0
    for a in asamalar:
        iscilik_dk += a["aktif_dakika"] if a["aktif_dakika"] is not None else a["sure_dakika"]
        if a["isil_islem_mi"] and a["agirlikli_ozgul_isi"] and a["isitilan_kutle_gram"]:
            kutle = a["isitilan_kutle_gram"] * porsiyon
            delta_t = a["hedef_sicaklik"] - a["baslangic_sicaklik"]
            joule = kutle * a["agirlikli_ozgul_isi"] * delta_t
            kwh = joule / 3_600_000.0 / a["verimlilik_orani"]
            birim_fiyat = (
                ayarlar["elektrik_birim_fiyat_eur_kwh"] if a["enerji_kaynagi"] == "elektrik"
                else ayarlar["dogalgaz_birim_fiyat_eur_kwh"]
            )
            enerji_eur += kwh * birim_fiyat
    iscilik_eur = (iscilik_dk / 60.0) * ayarlar["personel_saat_ucreti_eur"]
    return enerji_eur, iscilik_eur


mutfaklar_listesi = _mutfaklari_getir()
sol_mutfak, _bos_mutfak = st.columns([1, 3])
with sol_mutfak:
    mutfak_secimi = st.selectbox(
        "Mutfak", mutfaklar_listesi, format_func=lambda m: m["ad"],
    )
# Su an sadece Turk Mutfagi var, ama tarif sorgusu asagida mutfak_secimi["kod"]'a
# gore calisiyor -- ileride yeni bir mutfak (ör. Fransiz Mutfagi) eklenip
# kendi mutfak_kategorileri/receteler'i girildiginde, bolge butonlari ve
# tum akis otomatik olarak o mutfaga gore calisacak, ek kod degisikligi gerekmez.


@st.cache_data(ttl=3600)
def _tarif_kutuphanesini_getir(mutfak_kodu):
    mutfak = (
        supabase.table("mutfaklar").select("id").eq("kod", mutfak_kodu).single().execute()
    ).data
    kategoriler = (
        supabase.table("mutfak_kategorileri")
        .select("id, sira")
        .eq("mutfak_id", mutfak["id"])
        .execute()
    ).data
    grup_by_kategori = {k["id"]: k["sira"] for k in kategoriler}

    receteler = (
        supabase.table("receteler")
        .select("ad, mutfak_kategori_id, ozel_etiketler, mevsim_etiketi, bolge")
        .is_("isletme_id", "null")
        .execute()
    ).data

    tarifler = []
    for r in receteler:
        grup = grup_by_kategori.get(r["mutfak_kategori_id"])
        if grup is None:
            continue
        tarifler.append(
            {
                "ad": r["ad"],
                "grup": grup,
                "mevsim_etiketi": r["mevsim_etiketi"] or "yil_boyunca",
                "etiketler": r["ozel_etiketler"] or [],
                "bolge": r["bolge"] or "Genel",
            }
        )
    return tarifler


@st.cache_data(ttl=3600)
def _tarif_detaylarini_getir(isletme_id):
    """Global tariflerin porsiyon basi besin degeri, alerjen listesi ve
    (bu isletmenin kendi malzeme fiyatlariyla) maliyetini hesaplar.
    Maliyet isletmeye ozeldir cunku fiyatlar isletme_id'ye gore tutuluyor
    (malzeme_guncel_fiyat) -- global tariflerin kendi fiyati yok."""
    receteler = (
        supabase.table("receteler").select("id, ad").is_("isletme_id", "null").execute()
    ).data
    id_to_ad = {r["id"]: r["ad"] for r in receteler}

    malzeme_kalemleri = _sayfalayarak_getir(
        lambda: supabase.table("recete_malzemeleri")
        .select(
            "recete_id, malzeme_id, miktar_gram, "
            "malzemeler(ad, kalori, protein, yag, karbonhidrat, glisemik_indeks, fire_orani, "
            + ", ".join(_GENISLETILMIS_KOLONLAR) + ")"
        )
    )

    alerjen_kayitlari = _sayfalayarak_getir(
        lambda: supabase.table("malzeme_alerjen").select("malzeme_id, alerjenler(ad)")
    )
    alerjen_by_malzeme = {}
    for kayit in alerjen_kayitlari:
        ad = (kayit.get("alerjenler") or {}).get("ad")
        if ad:
            alerjen_by_malzeme.setdefault(kayit["malzeme_id"], set()).add(ad)

    fiyat_kayitlari = _sayfalayarak_getir(
        lambda: supabase.table("malzeme_guncel_fiyat")
        .select("malzeme_id, fiyat_eur")
        .eq("isletme_id", isletme_id)
    )
    fiyat_by_malzeme = {f["malzeme_id"]: f["fiyat_eur"] for f in fiyat_kayitlari}
    fiyat_verisi_var = len(fiyat_by_malzeme) > 0

    ham = {}
    for kalem in malzeme_kalemleri:
        ad = id_to_ad.get(kalem["recete_id"])
        if ad is None:
            continue  # baska bir isletmeye ait ozel tarif olabilir, atla
        m = kalem.get("malzemeler") or {}
        oran = kalem["miktar_gram"] / 100.0
        girdi = ham.setdefault(
            ad, {"kalori": 0.0, "protein": 0.0, "yag": 0.0, "karbonhidrat": 0.0,
                 "gi_agirlikli": 0.0, "gi_karb_toplam": 0.0, "maliyet_eur": 0.0,
                 "tam_fiyatli": True, "eksik_malzemeler": set(), "alerjenler": set(),
                 **{k: 0.0 for k in _GENISLETILMIS_KOLONLAR},
                 # ON DOKUZUNCU DUZELTME (13 Agustos 2026): bir tarifte
                 # HICBIR malzeme belirli bir besin ogesi icin veri
                 # icermiyorsa (ör. Vitamin B7 -- kataloğumuzda birçok
                 # malzemede hala eksik), toplamin "0" degil "bilinmiyor"
                 # (None) olmasi gerekiyor -- yoksa hedef kontrolu
                 # yanlislikla "0 < min" diyerek TUM ogunleri hedef disi
                 # isaretliyordu (kullanici 32 besin ogesinin hepsini
                 # secince bu hata ortaya cikti). Her genisletilmis kolon
                 # icin ayri bir "en az bir malzeme veri verdi mi" bayragi
                 # tutuyoruz.
                 **{f"{k}_var_mi": False for k in _GENISLETILMIS_KOLONLAR}}
        )
        girdi["kalori"] += (m.get("kalori") or 0) * oran
        girdi["protein"] += (m.get("protein") or 0) * oran
        girdi["yag"] += (m.get("yag") or 0) * oran
        karb = (m.get("karbonhidrat") or 0) * oran
        girdi["karbonhidrat"] += karb
        gi = m.get("glisemik_indeks")
        if gi is not None and karb > 0:
            girdi["gi_agirlikli"] += gi * karb
            girdi["gi_karb_toplam"] += karb
        for kolon in _GENISLETILMIS_KOLONLAR:
            deger = m.get(kolon)
            if deger is not None:
                girdi[kolon] += deger * oran
                girdi[f"{kolon}_var_mi"] = True

        malzeme_id = kalem["malzeme_id"]
        fiyat = fiyat_by_malzeme.get(malzeme_id)
        if fiyat is None:
            girdi["tam_fiyatli"] = False
            malzeme_adi = m.get("ad")
            if malzeme_adi:
                girdi["eksik_malzemeler"].add(malzeme_adi)
        else:
            # YETMIS YEDINCI DUZELTME (3 Eylul 2026): fire orani (soyma/
            # ayiklama kaybi) hesaba katiliyor -- 5_Tarif_Kutuphanesi.py
            # ve recete_guncel_maliyet SQL view ile AYNI formul, ucu
            # hesaplama yolu artik tutarli.
            fire_orani = m.get("fire_orani") or 0
            brut_miktar_gram = (
                kalem["miktar_gram"] / (1 - fire_orani)
                if fire_orani < 1 else kalem["miktar_gram"]
            )
            girdi["maliyet_eur"] += (brut_miktar_gram / 1000.0) * fiyat
        girdi["alerjenler"] |= alerjen_by_malzeme.get(malzeme_id, set())

    sonuc = {}
    for ad, v in ham.items():
        gi = (v["gi_agirlikli"] / v["gi_karb_toplam"]) if v["gi_karb_toplam"] > 0 else None
        sonuc[ad] = {
            "kalori": v["kalori"], "protein": v["protein"],
            "yag": v["yag"], "karbonhidrat": v["karbonhidrat"], "gi": gi,
            "maliyet_eur": v["maliyet_eur"], "tam_fiyatli": v["tam_fiyatli"],
            "eksik_malzemeler": v["eksik_malzemeler"], "alerjenler": v["alerjenler"],
            **{k: (v[k] if v[f"{k}_var_mi"] else None) for k in _GENISLETILMIS_KOLONLAR},
        }
    return sonuc, fiyat_verisi_var


def _ogun_toplami(tarif_adlari, detay):
    toplam = {"kalori": 0.0, "protein": 0.0, "yag": 0.0, "karbonhidrat": 0.0, "maliyet_eur": 0.0,
              **{k: 0.0 for k in _GENISLETILMIS_KOLONLAR}}
    # Ayni "veri var mi" ayrimi burada da gerekli -- meal'deki HICBIR
    # tarif belirli bir besin ogesi icin veri tasimiyorsa, o ogenin
    # ogun toplami "0" degil None olmali (bkz. yukaridaki not).
    genisletilmis_var_mi = {k: False for k in _GENISLETILMIS_KOLONLAR}
    gi_agirlikli = 0.0
    gi_karb_toplam = 0.0
    tam_fiyatli = True
    eksik_malzemeler = set()
    alerjenler = set()
    for ad in tarif_adlari:
        b = detay.get(ad)
        if not b:
            continue
        toplam["kalori"] += b["kalori"]
        toplam["protein"] += b["protein"]
        toplam["yag"] += b["yag"]
        toplam["karbonhidrat"] += b["karbonhidrat"]
        toplam["maliyet_eur"] += b["maliyet_eur"]
        tam_fiyatli = tam_fiyatli and b["tam_fiyatli"]
        eksik_malzemeler |= b["eksik_malzemeler"]
        alerjenler |= b["alerjenler"]
        if b["gi"] is not None and b["karbonhidrat"] > 0:
            gi_agirlikli += b["gi"] * b["karbonhidrat"]
            gi_karb_toplam += b["karbonhidrat"]
        for kolon in _GENISLETILMIS_KOLONLAR:
            deger = b.get(kolon)
            if deger is not None:
                toplam[kolon] += deger
                genisletilmis_var_mi[kolon] = True
    for kolon in _GENISLETILMIS_KOLONLAR:
        if not genisletilmis_var_mi[kolon]:
            toplam[kolon] = None
    toplam["gi"] = round(gi_agirlikli / gi_karb_toplam) if gi_karb_toplam > 0 else None
    toplam["tam_fiyatli"] = tam_fiyatli
    toplam["eksik_malzemeler"] = eksik_malzemeler
    toplam["alerjenler"] = alerjenler
    return toplam


@st.cache_data(ttl=3600)
def _isletme_receteler_ve_detay_getir(isletme_id):
    """Isletmenin kendi ozel receteleri (1_Receteler.py'de olusturulan)
    -- Yillik Menu'ye ISTEGE BAGLI olarak eklenebilir (bkz. asagidaki
    'kendi menu' butonu). Kutuphane tarifleriyle AYNI sekle
    (ad/grup/mevsim_etiketi/etiketler/bolge + besin/maliyet detayi)
    donusturulur ki uretim_algoritmasi.py'ye hic dokunmadan ayni
    fonksiyona (hafta_olustur) verilebilsin.

    Kategori -> anayasa grubu eslesmesi (6 Agustos 2026, kullaniciyla
    netlestirildi):
      ana_yemek -> 1 (Ana Yemek), corba -> 2 (Yardimci Yemek),
      salata/tatli -> 3 (Tamamlayici),
      icecek/baslangic/pizza/burger -> 4 (ISTEGE BAGLI Fast Food yuvasi
      -- anayasa madde 8'in ZORUNLU 3'lusune DAHIL DEGIL, bkz.
      uretim_algoritmasi.py._fast_food_sec).

    Bolge: ozel receteler bolge bilgisi TASIMAZ ve bolge filtresinden
    MUAF tutulur (kullanicinin acik talebi, 6 Agustos 2026) -- "bolge"
    alani sentinel bir deger (__isletme__) tasir ama cagiran kod bunlari
    zaten bolge filtresi UYGULANDIKTAN SONRA havuza ekler, o yuzden bu
    deger hicbir filtrede kullanilmaz.

    Maliyet hesabi kutuphane fonksiyonuyla (_tarif_detaylarini_getir)
    AYNI dogrudan malzeme_guncel_fiyat yontemini kullanir (1_Receteler.py
    'deki recete_guncel_maliyet VIEW'inden FARKLI bir yol -- iki ayri
    yontemin sonucu ayni olmali ama bagimsiz olarak hesaplaniyor)."""
    KATEGORI_GRUP = {
        "ana_yemek": 1, "corba": 2, "salata": 3, "tatli": 3,
        "icecek": 4, "baslangic": 4, "pizza": 4, "burger": 4,
    }

    receteler = (
        supabase.table("receteler")
        .select("id, ad, kategori, porsiyon_sayisi")
        .eq("isletme_id", isletme_id)
        .execute()
    ).data or []
    if not receteler:
        return [], {}, False

    id_to_ad = {r["id"]: r["ad"] for r in receteler}
    porsiyon_by_id = {r["id"]: (r["porsiyon_sayisi"] or 1) for r in receteler}

    malzeme_kalemleri = _sayfalayarak_getir(
        lambda: supabase.table("recete_malzemeleri")
        .select(
            "recete_id, malzeme_id, miktar_gram, "
            "malzemeler(ad, kalori, protein, yag, karbonhidrat, glisemik_indeks, fire_orani, "
            + ", ".join(_GENISLETILMIS_KOLONLAR) + ")"
        )
        .in_("recete_id", list(id_to_ad.keys()))
    ) or []

    alerjen_kayitlari = _sayfalayarak_getir(
        lambda: supabase.table("malzeme_alerjen").select("malzeme_id, alerjenler(ad)")
    ) or []
    alerjen_by_malzeme = {}
    for kayit in alerjen_kayitlari:
        ad = (kayit.get("alerjenler") or {}).get("ad")
        if ad:
            alerjen_by_malzeme.setdefault(kayit["malzeme_id"], set()).add(ad)

    fiyat_kayitlari = _sayfalayarak_getir(
        lambda: supabase.table("malzeme_guncel_fiyat")
        .select("malzeme_id, fiyat_eur")
        .eq("isletme_id", isletme_id)
    ) or []
    fiyat_by_malzeme = {f["malzeme_id"]: f["fiyat_eur"] for f in fiyat_kayitlari}
    fiyat_verisi_var = len(fiyat_by_malzeme) > 0

    ham = {}
    for kalem in malzeme_kalemleri:
        recete_id = kalem["recete_id"]
        if recete_id not in id_to_ad:
            continue
        m = kalem.get("malzemeler") or {}
        oran = kalem["miktar_gram"] / 100.0
        girdi = ham.setdefault(
            recete_id, {"kalori": 0.0, "protein": 0.0, "yag": 0.0, "karbonhidrat": 0.0,
                        "gi_agirlikli": 0.0, "gi_karb_toplam": 0.0, "maliyet_eur": 0.0,
                        "tam_fiyatli": True, "eksik_malzemeler": set(), "alerjenler": set(),
                        **{k: 0.0 for k in _GENISLETILMIS_KOLONLAR},
                        **{f"{k}_var_mi": False for k in _GENISLETILMIS_KOLONLAR}}
        )
        girdi["kalori"] += (m.get("kalori") or 0) * oran
        girdi["protein"] += (m.get("protein") or 0) * oran
        girdi["yag"] += (m.get("yag") or 0) * oran
        karb = (m.get("karbonhidrat") or 0) * oran
        girdi["karbonhidrat"] += karb
        gi = m.get("glisemik_indeks")
        if gi is not None and karb > 0:
            girdi["gi_agirlikli"] += gi * karb
            girdi["gi_karb_toplam"] += karb
        for kolon in _GENISLETILMIS_KOLONLAR:
            deger = m.get(kolon)
            if deger is not None:
                girdi[kolon] += deger * oran
                girdi[f"{kolon}_var_mi"] = True

        malzeme_id = kalem["malzeme_id"]
        fiyat = fiyat_by_malzeme.get(malzeme_id)
        if fiyat is None:
            girdi["tam_fiyatli"] = False
            malzeme_adi = m.get("ad")
            if malzeme_adi:
                girdi["eksik_malzemeler"].add(malzeme_adi)
        else:
            # YETMIS YEDINCI DUZELTME (3 Eylul 2026): fire orani (soyma/
            # ayiklama kaybi) hesaba katiliyor -- 5_Tarif_Kutuphanesi.py
            # ve recete_guncel_maliyet SQL view ile AYNI formul, ucu
            # hesaplama yolu artik tutarli.
            fire_orani = m.get("fire_orani") or 0
            brut_miktar_gram = (
                kalem["miktar_gram"] / (1 - fire_orani)
                if fire_orani < 1 else kalem["miktar_gram"]
            )
            girdi["maliyet_eur"] += (brut_miktar_gram / 1000.0) * fiyat
        girdi["alerjenler"] |= alerjen_by_malzeme.get(malzeme_id, set())

    tarif_listesi = []
    detay = {}
    for r in receteler:
        grup = KATEGORI_GRUP.get(r["kategori"])
        if grup is None:
            continue  # bilinmeyen/yeni bir kategori -- sessizce atla
        tarif_listesi.append({
            "ad": r["ad"], "grup": grup, "mevsim_etiketi": "yil_boyunca",
            "etiketler": [], "bolge": "__isletme__",
        })

        v = ham.get(r["id"])
        porsiyon = porsiyon_by_id[r["id"]]
        if v is None:
            detay[r["ad"]] = {
                "kalori": 0.0, "protein": 0.0, "yag": 0.0, "karbonhidrat": 0.0, "gi": None,
                "maliyet_eur": 0.0, "tam_fiyatli": True, "eksik_malzemeler": set(), "alerjenler": set(),
                **{k: None for k in _GENISLETILMIS_KOLONLAR},
            }
            continue
        gi = (v["gi_agirlikli"] / v["gi_karb_toplam"]) if v["gi_karb_toplam"] > 0 else None
        detay[r["ad"]] = {
            "kalori": v["kalori"] / porsiyon, "protein": v["protein"] / porsiyon,
            "yag": v["yag"] / porsiyon, "karbonhidrat": v["karbonhidrat"] / porsiyon, "gi": gi,
            "maliyet_eur": v["maliyet_eur"] / porsiyon, "tam_fiyatli": v["tam_fiyatli"],
            "eksik_malzemeler": v["eksik_malzemeler"], "alerjenler": v["alerjenler"],
            **{k: (v[k] / porsiyon if v[f"{k}_var_mi"] else None) for k in _GENISLETILMIS_KOLONLAR},
        }

    return tarif_listesi, detay, fiyat_verisi_var


tarifler = _tarif_kutuphanesini_getir(mutfak_secimi["kod"])

if not tarifler:
    st.warning(
        "Global tarif kütüphanesi boş görünüyor. Önce `yukle_tarifler.py` "
        "ile 74 tarifin Supabase'e yüklendiğinden emin ol."
    )
    st.stop()

st.caption(f"Kütüphanede {len(tarifler)} tarif bulundu.")

KISA_BOLGE_ADI = {
    "Genel": "Klasik",
    "Doğu Anadolu": "Doğu",
    "Güneydoğu Anadolu": "Güneydoğu",
}

BOLGE_SIRASI = ["Marmara", "Ege", "Akdeniz", "Karadeniz", "İç Anadolu", "Doğu Anadolu", "Güneydoğu Anadolu"]
mevcut_diger_bolgeler = {t["bolge"] for t in tarifler} - {"Genel"}
diger_bolgeler = [b for b in BOLGE_SIRASI if b in mevcut_diger_bolgeler]
diger_bolgeler += sorted(mevcut_diger_bolgeler - set(BOLGE_SIRASI))  # BOLGE_SIRASI'nda olmayan yeni bolgeler sona eklenir
bolgeler_mevcut = (["Genel"] if any(t["bolge"] == "Genel" for t in tarifler) else []) + diger_bolgeler

if "secili_bolgeler_set" not in st.session_state:
    st.session_state.secili_bolgeler_set = set()  # bos = hicbir kisit yok, tumu kullanilir
if "kendi_menu_dahil" not in st.session_state:
    st.session_state.kendi_menu_dahil = False

isletme_bilgi = (
    supabase.table("isletmeler").select("kisaltma").eq("id", st.session_state.isletme_id).single().execute()
).data
# OTUZ DOKUZUNCU DUZELTME (30 Agustos 2026): kullanici talebiyle -- bu
# buton artik isletmenin TAM adi degil, KISALTILMIS adi (bkz. Abonelik
# sayfasi, 75 numarali migration) ile etiketleniyor -- satirdaki diger
# butonlarla (kisa bolge adlari) boy/stil olarak tutarli olsun diye.
# Kisaltma girilmemisse jenerik "ÖZEL" kullanilir.
isletme_kisaltma = ((isletme_bilgi or {}).get("kisaltma") or "").strip()
isletme_adi = isletme_kisaltma or "ÖZEL"

st.markdown("**Bölge (mutfak)**")
st.caption(
    "Hiçbiri seçili değilken tüm bölgeler kullanılır. Bir bölgeye tıklamak "
    f"SADECE onu etkinleştirir. \"{isletme_adi}\" butonu, işletmenin kendi "
    "özel reçetelerini (bölge seçiminden bağımsız, her zaman) menüye dahil eder."
)
kolonlar = st.columns(len(bolgeler_mevcut) + 1)
for kolon, bolge in zip(kolonlar[:-1], bolgeler_mevcut):
    secili = bolge in st.session_state.secili_bolgeler_set
    etiket = KISA_BOLGE_ADI.get(bolge, bolge)
    if kolon.button(
        etiket, key=f"bolge_buton_{bolge}", use_container_width=True,
        type="primary" if secili else "secondary",
    ):
        if secili:
            st.session_state.secili_bolgeler_set.discard(bolge)
        else:
            st.session_state.secili_bolgeler_set.add(bolge)
        st.rerun()

with kolonlar[-1]:
    if st.button(
        isletme_adi, key="kendi_menu_buton", use_container_width=True,
        type="primary" if st.session_state.kendi_menu_dahil else "secondary",
    ):
        st.session_state.kendi_menu_dahil = not st.session_state.kendi_menu_dahil
        st.rerun()

secili_bolgeler = st.session_state.secili_bolgeler_set

if secili_bolgeler:
    tarifler = [t for t in tarifler if t["bolge"] in secili_bolgeler]
# secili_bolgeler bossa (hicbir buton tiklanmamissa) hicbir filtre uygulanmaz, tum bolgeler kullanilir

# "Kendi menum" -- BOLGE FILTRESI UYGULANDIKTAN SONRA, filtreden MUAF
# olarak havuza ekleniyor (kullanicinin acik talebi).
detay_ozel = {}
fiyat_ozel_var = False
if st.session_state.kendi_menu_dahil:
    ozel_tarifler, detay_ozel, fiyat_ozel_var = _isletme_receteler_ve_detay_getir(
        st.session_state.isletme_id
    )
    if not ozel_tarifler:
        st.caption(
            f"\"{isletme_adi}\" için henüz Ana Yemek/Çorba/Salata/Tatlı/"
            "İçecek/Başlangıç/Pizza/Burger kategorisinde bir reçete yok."
        )
    tarifler = tarifler + ozel_tarifler

if not tarifler:
    st.warning("Seçtiğin bölge(ler)de hiç tarif bulunamadı.")
    st.stop()

if secili_bolgeler:
    st.caption(f"Seçili bölge(ler)de {len(tarifler)} tarif kullanılacak.")
else:
    st.caption(f"Hiçbir bölge seçilmedi, tüm {len(tarifler)} tarif kullanılacak.")


detay, fiyat_verisi_var = _tarif_detaylarini_getir(st.session_state.isletme_id)
if detay_ozel:
    detay = {**detay, **detay_ozel}
    fiyat_verisi_var = fiyat_verisi_var or fiyat_ozel_var
if not fiyat_verisi_var:
    st.caption(
        "Bu işletme için henüz malzeme fiyatı girilmemiş — maliyet "
        "sütunu bu yüzden hesaplanamıyor (\"-\" gösterilecek)."
    )

# Uretim algoritmasi besin hedefi kontrolu icin HER alani (kalori/protein/
# yag/karbonhidrat/gi + 27 genisletilmis besin ogesi) tarife ekliyoruz
# (detay'dan -- zaten hesaplanmisti). Hangilerinin GERCEKTEN hedeflenecegi
# asagidaki multiselect ile secilir, ama _hedefte_mi kontrolunun her alana
# erisebilmesi icin hepsi burada tasiniyor.
tarifler_zengin = []
for t in tarifler:
    b = detay.get(t["ad"], {})
    t2 = dict(t)
    t2["kalori"] = b.get("kalori")
    t2["protein"] = b.get("protein")
    t2["yag"] = b.get("yag")
    t2["karbonhidrat"] = b.get("karbonhidrat")
    t2["gi"] = b.get("gi")
    for kolon in _GENISLETILMIS_KOLONLAR:
        t2[kolon] = b.get(kolon)
    tarifler_zengin.append(t2)

sol, sag, sag2, sag3 = st.columns([1, 1, 1.4, 1.6])
with sol:
    yil_secimi = st.number_input(
        "Yıl", min_value=2024, max_value=2035, value=datetime.date.today().year, step=1,
    )
with sag:
    ay_secimi = st.selectbox("Ay", AYLAR_SIRALI)
with sag2:
    # SEKSEN BIRINCI DUZELTME (3 Eylul 2026): porsiyon profili secimi
    # pop-up'tan buraya (Yıl/Ay'ın yanina) tasindi -- Bahri'nin talebi,
    # boylece her pop-up'ta ayri ayri secmek yerine bir kez secilip TUM
    # gunler icin gecerli oluyor. Secilen deger session_state'e yazilir,
    # pop-up (_gun_popup_govdesini_ciz) oradan okur.
    #
    # SEKSEN IKINCI DUZELTME (4 Eylul 2026): profil artik sadece
    # porsiyon degil, kendi besin hedeflerini de tasiyabiliyor (bkz. 80
    # numarali migration, Abonelik sayfasindaki "Profil Basina Besin
    # Hedefleri" bolumu). Profil DEGISTIRILDIGINDE, o profilin kayitli
    # hedefleri asagidaki "Ogun basina besin hedefi" widget'larina
    # OTOMATIK yukleniyor -- kullanici hala isterse o an icin
    # degistirebilir, ama baslangic noktasi artik dogru kurumdan geliyor.
    _porsiyon_profilleri_sayfa = (
        supabase.table("isletme_porsiyon_profilleri")
        .select("id, ad, porsiyon_sayisi, hedefler")
        .eq("isletme_id", st.session_state.isletme_id)
        .order("sira")
        .execute()
    ).data or []
    # YUZ ONBIRINCI DUZELTME (5 Eylul 2026): Bahri'nin talebi -- bu
    # listedeki TUM secenekler onun Abonelik sayfasinda olusturdugu
    # GERCEK profillerdi. Eger BUNLARIN HEPSININ zaten kayitli hedefi
    # varsa, "Ogun basina besin hedefi" ad-hoc/manuel giris arayuzune
    # HICBIR YOLLA erisemiyordu. Listenin sonuna, hedefi HIC olmayan
    # (bu yuzden asagida otomatik olarak manuel giris arayuzunu acan)
    # sentetik bir "Boş Profil" secenegi eklendi -- Abonelik'e
    # KAYDEDILMEDEN, bu SADECE bu uretim icin gecici bir hedef girmeyi
    # sagliyor.
    _porsiyon_profilleri_sayfa = _porsiyon_profilleri_sayfa + [
        {"id": None, "ad": "Boş Profil (özel/geçici hedef)", "porsiyon_sayisi": 10, "hedefler": None}
    ]
    _profil_etiketleri_sayfa = [f"{p['ad']} ({p['porsiyon_sayisi']} porsiyon)" for p in _porsiyon_profilleri_sayfa]
    _sayfa_secili_index = st.selectbox(
        "Maliyet hesabı için porsiyon profili",
        options=range(len(_profil_etiketleri_sayfa)),
        format_func=lambda i: _profil_etiketleri_sayfa[i],
        key="sayfa_porsiyon_profili_secimi",
        help="Profilleri eklemek/düzenlemek için Abonelik sayfasındaki "
             "\"Porsiyon Profilleri\" bölümüne bak.",
    )
    _secili_sayfa_profili = _porsiyon_profilleri_sayfa[_sayfa_secili_index]
    st.session_state["secili_porsiyon_profil_id"] = _secili_sayfa_profili["id"]
    st.session_state["secili_porsiyon_sayisi"] = _secili_sayfa_profili["porsiyon_sayisi"]

with sag3:
    # DOKSAN UCUNCU DUZELTME (4 Eylul 2026): Bahri'nin talebi -- daha
    # once KAYDEDILMIS aylik menuleri (bkz. 81 numarali migration),
    # secili PROFILE gore listeleyip geri YUKLEYEBILME. Secilince
    # yillik_menu_aylik/yillik_menu_hedefler dogrudan degistirilip
    # rerun ediliyor -- boylece ayni sayfanin ALTINDAKI goruntuleme/
    # Excel/kaydet mantigi HICBIR DEGISIKLIK gerektirmeden calisir.
    _kayitli_menuler = []
    if _secili_sayfa_profili["id"]:
        _kayitli_menuler = (
            supabase.table("kayitli_aylik_menuler")
            .select("id, yil, ay, menu_verisi")
            .eq("isletme_id", st.session_state.isletme_id)
            .eq("porsiyon_profil_id", _secili_sayfa_profili["id"])
            .order("yil", desc=True)
            .execute()
        ).data or []

    if _kayitli_menuler:
        _kayitli_etiketler = ["— Seç —"] + [f"{m['ay']} {m['yil']}" for m in _kayitli_menuler]
        _kayitli_secim = st.selectbox(
            "Bu Profil için Kayıtlı Aylık Menüler",
            options=_kayitli_etiketler,
            key=f"kayitli_menu_secimi_{_secili_sayfa_profili['id']}",
        )
        if _kayitli_secim != "— Seç —":
            _secilen_kayit = _kayitli_menuler[_kayitli_etiketler.index(_kayitli_secim) - 1]
            _onceki_yuklenen = st.session_state.get("_yuklenen_kayitli_menu_id")
            if _onceki_yuklenen != _secilen_kayit["id"]:
                st.session_state["_yuklenen_kayitli_menu_id"] = _secilen_kayit["id"]
                st.session_state["yillik_menu_aylik"] = {
                    "ay": _secilen_kayit["ay"], "yil": _secilen_kayit["yil"],
                    "haftalar": _secilen_kayit["menu_verisi"]["haftalar"], "gecmis_ay_mi": False,
                }
                # NOT: buradaki "hedefler" (sayfa degiskeni) HENUZ
                # tanimlanmadi -- "Ogun basina besin hedefi" bolumu bu
                # noktadan SONRA calisiyor. Onun yerine profilin HAM
                # verisini (ayni kaynak) dogrudan kullaniyoruz.
                st.session_state["yillik_menu_hedefler"] = _secili_sayfa_profili.get("hedefler")
                st.rerun()
    else:
        st.caption("Bu profil için henüz kaydedilmiş aylık menü yok.")

    # Profil DEGISTI mi VEYA ayni profilin KAYITLI HEDEFLERI DEGISTI mi
    # kontrol et -- hedef widget'larinin session_state'ini bu profilin
    # kayitli hedefleriyle ONCEDEN doldur (widget'lar HENUZ olusturulmadi,
    # asagida olusturulacak -- Streamlit'te bir widget'in key'i
    # session_state'te zaten varsa value= parametresi yok sayilir, o
    # yuzden bu ON-DOLDURMA'nin widget'lardan ONCE calismasi sart).
    #
    # SEKSEN UCUNCU DUZELTME (4 Eylul 2026): SADECE profil ID'sini
    # karsilastirmak YETERSIZDI -- Bahri, Abonelik'te ZATEN SECILI olan
    # bir profilin (ör. "EV") hedeflerini degistirip kaydettiginde,
    # Aylik Menu'ye donup AYNI profili (yine "EV") secince degisiklikler
    # YUKLENMEDI -- cunku ID degismemisti, guard "hicbir sey degismedi"
    # sanip atladi. Simdi ID + hedeflerin kendisi (JSON) birlikte bir
    # "imza" olusturuyor -- profilin ALTINDAKI VERI degisse bile
    # (ID ayni kalsa dahi) yeniden yukleme tetikleniyor.
    _hedef_imzasi = (_secili_sayfa_profili["id"], json.dumps(_secili_sayfa_profili.get("hedefler"), sort_keys=True))
    if st.session_state.get("_onceki_profil_imza_aylik_menu") != _hedef_imzasi:
        st.session_state["_onceki_profil_imza_aylik_menu"] = _hedef_imzasi
        _profil_hedefleri = _secili_sayfa_profili.get("hedefler") or {}
        if _profil_hedefleri:
            st.session_state["besin_hedefi_kullan"] = True
            st.session_state["yillik_menu_secili_besin_anahtarlari"] = kanonik_sirala(
                {anahtar for ogun in _profil_hedefleri.values() for anahtar in ogun}
            )
            for _ogun_adi, _degerler in _profil_hedefleri.items():
                for _anahtar, _aralik in _degerler.items():
                    st.session_state[f"{_ogun_adi}_{_anahtar}_alt"] = float(_aralik[0])
                    st.session_state[f"{_ogun_adi}_{_anahtar}_ust"] = float(_aralik[1])

# SEKSEN BESINCI DUZELTME (4 Eylul 2026): Bahri'nin talebi -- secili
# profilin ZATEN kayitli besin hedefleri varsa, "Ogun basina besin
# hedefi uygula" sorusu/arayuzu HIC gosterilmiyor, hedefler sessizce
# otomatik uygulaniyor. Sadece profilin HIC hedefi yoksa (ör. "Standart"
# gibi genel amacli bir profil), eskisi gibi manuel/gecici bir hedef
# girme secenegi sunuluyor.
_profil_kayitli_hedefleri = _secili_sayfa_profili.get("hedefler") or {}

hedefler = None
if _profil_kayitli_hedefleri:
    st.info(
        f"\"{_secili_sayfa_profili['ad']}\" profili için besin hedefleri "
        "zaten tanımlı, bu üretimde otomatik uygulanacak. Değiştirmek "
        "istersen Abonelik sayfasındaki \"Profil Başına Besin "
        "Hedefleri\" bölümüne bak."
    )
    hedefler = _profil_kayitli_hedefleri
else:
    besin_hedefi_kullan = st.checkbox(
        "Öğün başına besin hedefi uygula (opsiyonel)", key="besin_hedefi_kullan",
    )

    if besin_hedefi_kullan:
        st.caption(
            "Önce hangi besin değerlerini hedeflemek istediğini seç (kalori "
            "gibi temel değerler varsayılan olarak seçili) — sadece seçtiklerin "
            "için aşağıda min/maks aralığı gösterilecek."
        )
        secili_besin_anahtarlari = st.multiselect(
            "Hedeflenecek besin değerleri",
            options=[anahtar for anahtar, *_ in TUM_BESIN_ALANLARI],
            default=["kalori", "protein", "yag", "karbonhidrat", "gi"],
            format_func=lambda a: BESIN_ETIKET[a],
            key="yillik_menu_secili_besin_anahtarlari",
        )
        hedefler = {}
        for ogun_adi in ("Öğle", "Akşam"):
            with st.expander(f"{ogun_adi} hedefleri", expanded=False):
                hedefler[ogun_adi] = {}
                if not secili_besin_anahtarlari:
                    st.caption("Yukarıdan en az bir besin değeri seçmelisin.")
                for anahtar in secili_besin_anahtarlari:
                    etiket = BESIN_ETIKET[anahtar]
                    # (float(...) burada bilinçli bir guvenlik agi: TUM_BESIN_ALANLARI'na
                    # ileride eklenecek bir satirda min/maks/varsayilan turleri
                    # yanlislikla karisik (int+float) yazilirsa bile, number_input
                    # yine de tek tip float alacak -- StreamlitMixedNumericTypesError
                    # bir daha tekrarlanmasin diye.)
                    minv, maxv, def_alt, def_ust = (float(x) for x in BESIN_ARALIK[anahtar])
                    c1, c2 = st.columns(2)
                    with c1:
                        alt = st.number_input(
                            f"{etiket} — min", min_value=minv, max_value=maxv,
                            value=def_alt, key=f"{ogun_adi}_{anahtar}_alt",
                        )
                    with c2:
                        ust = st.number_input(
                            f"{etiket} — maks", min_value=minv, max_value=maxv,
                            value=def_ust, key=f"{ogun_adi}_{anahtar}_ust",
                        )
                    hedefler[ogun_adi][anahtar] = (alt, ust)

if hedefler and secili_bolgeler and len(tarifler) < 60:
    st.caption(
        "Dikkat: dar bir bölge seçiliyken besin hedefi de uygularsan, küçük "
        "havuzda hedefe uyan tarif sayısı çok azalabilir ve menü tek bir "
        "yemeğe kilitlenebilir. Çeşitlilik azsa hedef aralığını genişletmeyi "
        "veya daha fazla bölge seçmeyi dene."
    )

if st.button("Ay için menü üret", type="primary"):
    # SEKSEN BESINCI DUZELTME (4 Eylul 2026): uretim uzun surebiliyor
    # (bircok hafta, her biri icin cok sayida deneme) -- Bahri, bu
    # sure boyunca ekranin "donmus" gibi gorunmesinden rahatsiz oldu.
    # st.spinner ile aciken bir "uretiliyor" gostergesi eklendi.
    with st.spinner("Menü üretiliyor, lütfen bekleyin..."):
        gercek_haftalar = _ay_gercek_haftalari(yil_secimi, ay_secimi)
        haftalar = []
        for hafta_tarihleri in gercek_haftalar:
            # Deterministik tohum: haftanin Pazartesi gununun takvim sirasina
            # (toordinal) gore -- ayni yil+hafta her zaman ayni sonucu verir,
            # eski "ay_index*10+hafta_no" semasi gercek haftalar ay sinirini
            # astigi icin artik anlamli degildi.
            tohum = hafta_tarihleri[0].toordinal()
            rastgele = random.Random(tohum)
            gun_mevsimleri = [_tarih_mevsimi(t) for t in hafta_tarihleri]
            hafta = hafta_olustur(
                tarifler_zengin, None, rastgele, hedefler=hedefler,
                gun_mevsimleri=gun_mevsimleri,
            )
            # Ekranda gercek tarih/hafta gunu adi gosterebilmek icin, o gunun
            # gercek datetime.date'ini de tasiyoruz.
            for gun, tarih in zip(hafta, hafta_tarihleri):
                gun["tarih"] = tarih
            haftalar.append(hafta)
    # YIRMI IKINCI DUZELTME (13 Agustos 2026, Oturum 11): kullanicinin
    # "gecmis bir tarih icin uretilen menu, gercek servis kaydiyla
    # karistirilabilir" endisesi uzerine -- uretimi ENGELLEMIYORUZ (test/
    # karsilastirma gibi mesru kullanim senaryolari var), sadece secilen
    # ayin GERCEK son gunu bugunden ONCEYSE (yani tum ay gecmiste kaldiysa)
    # ekranda acik bir uyari notu gosteriyoruz -- kullaniciyla mutabik
    # kalinan cozum bu.
    _ay_son_gunu = gercek_haftalar[-1][-1] if gercek_haftalar else None
    _gecmis_ay_mi = _ay_son_gunu is not None and _ay_son_gunu < datetime.date.today()
    st.session_state["yillik_menu_aylik"] = {
        "ay": ay_secimi, "yil": yil_secimi, "haftalar": haftalar, "gecmis_ay_mi": _gecmis_ay_mi,
    }
    st.session_state["yillik_menu_hedefler"] = hedefler

RENKLER = {1: "#D85A30", 2: "#639922", 3: "#1D9E75"}


# (TEMEL_5 artik dosyanin en basinda tanimli -- uretim cagrisindan
# ONCE gelmesi gerekiyordu, bkz. DOKSAN ALTINCI DUZELTME.)


def _hedef_disi_liste_metni(kayitlar, yil_secimi, ay_secimi):
    """YUZUNCU DUZELTME (4 Eylul 2026): Bahri'nin talebi -- tarih formati
    ISO (2026-12-01) DEGIL Turkce olsun, VE ardisik ayni-ay/yil gunler
    TEK BIR onek altinda gruplansin (ör. "2027 Ocak 1 (Öğle), 2 (Öğle),
    3 (Akşam)").

    ON BIRINCI DUZELTME (5 Eylul 2026, DUZELTME): ilk versiyon, SEÇILI
    ay icin HICBIR ZAMAN onek gostermiyordu (Bahri'nin "zaten secili,
    tekrar yazmaya gerek yok" talebine gore) -- ama bu, bir ONCEKI
    (secili OLMAYAN, ör. Aralık) baglamdan SECILI aya (ör. Ocak) GECIS
    yapildiginda hicbir isaret olmadan "...31 (Akşam), 1 (Akşam), 2..."
    seklinde devam ediyordu -- okuyucu "1"in Ocak'in 1'i mi yoksa
    Aralık'ta bir yerde mi oldugunu ANLAYAMIYORDU. Duzeltme: artik HER
    baglam degisiminde (SECILI aya gecis DAHIL) bir kez ay adi
    gosteriliyor, SADECE ayni baglam icinde ART ARDA gelen gunlerde
    onek atlaniyor -- boylece hem "zaten secili ayda tekrar tekrar ay
    yazma" hem "gecis noktasinda belirsizlik" sorunlari birlikte
    cozuluyor.

    kayitlar: [(tarih:date|None, ogun_adi:str), ...]"""
    parcalar = []
    onceki_baglam = None
    for tarih, ogun_adi in kayitlar:
        if tarih is None:
            parcalar.append(f"(tarihsiz, {ogun_adi})")
            onceki_baglam = "BILINMIYOR"
            continue
        baglam = (tarih.year, tarih.month)
        if baglam != onceki_baglam:
            ay_adi = AYLAR_SIRALI[tarih.month - 1]
            onek = f"{tarih.year} {ay_adi} " if tarih.year != yil_secimi else f"{ay_adi} "
        else:
            onek = ""
        parcalar.append(f"{onek}{tarih.day} ({ogun_adi})")
        onceki_baglam = baglam
    return ", ".join(parcalar)


def _haftalik_ortalama(ogun_adi, anahtar, hafta, detay):
    """Bir haftadaki (7 gun) TUM gunlerin, belirtilen ogun+besin ogesi
    icin ORTALAMASINI hesaplar. SEKSEN DOKUZUNCU DUZELTME (4 Eylul
    2026): TEMEL_5 disindaki (vitamin/mineral/vb., 22 oge) besin
    hedefleri artik TEK OGUN bazinda degil, HAFTALIK ORTALAMA bazinda
    kontrol ediliyor -- Bahri'nin de belirttigi gibi hicbir gercek
    yemek, 27 farkli besin ogesini AYNI ANDA dar bir araliga oturtamiyor
    (gercek diyetisyenler de mikro besinleri haftalik/aylik dengeler,
    tek ogun bazinda degil) -- 241 tariflik kutuphaneyle bunu HER
    OGUNDE tutturmaya calismak matematiksel olarak asiri kisitlayiciydi."""
    degerler = []
    for gun2 in hafta:
        tarif_adlari = (gun2.get("ogunler") or {}).get(ogun_adi)
        if not tarif_adlari:
            continue
        t2 = _ogun_toplami(tarif_adlari, detay)
        deger = t2.get(anahtar)
        if deger is not None:
            degerler.append(deger)
    if not degerler:
        return None
    return sum(degerler) / len(degerler)


def _hedefte_mi(ogun_adi, t, hedefler, hafta=None, detay=None):
    """TEMEL_5 (kalori/protein/yag/karbonhidrat/gi) HALA TEK OGUN (gun)
    bazinda sikica kontrol ediliyor -- bunlar ana/hemen-belirgin
    degerler, gun gun sapma onemli. TEMEL_5 DISINDAKI ogeler ise (hafta
    ve detay parametreleri verilmisse) HAFTALIK ORTALAMA uzerinden
    kontrol ediliyor (bkz. _haftalik_ortalama). hafta/detay verilmezse
    (ör. eski cagri yerleri, ya da haftalik baglam mevcut degilse) TUM
    anahtarlar eskisi gibi TEK OGUN bazinda kontrol edilir -- geriye
    donuk uyumluluk.

    SEKSEN DOKUZUNCU DUZELTME (4 Eylul 2026): donus degeri artik
    (True/False/None, basarisiz_olan_anahtar_listesi) ikilisi --
    Bahri'nin talebi: "Hedef dışı" yazisinin yaninda HANGI besin
    ogesinin hedef disi oldugu da gosterilsin."""
    if not hedefler or ogun_adi not in hedefler:
        return None, []
    basarisiz = []
    for anahtar, (alt, ust) in hedefler[ogun_adi].items():
        if anahtar in TEMEL_5 or hafta is None or detay is None:
            deger = t.get(anahtar)
        else:
            deger = _haftalik_ortalama(ogun_adi, anahtar, hafta, detay)
        if deger is None:
            continue
        if not (alt <= deger <= ust):
            basarisiz.append(anahtar)
    if basarisiz:
        return False, basarisiz
    return True, []


def _tablo_stilini_uygula():
    """YUZ ALTINCI DUZELTME (5 Eylul 2026): artik HER "satir" (tarih,
    gun adi, her Ogle/Aksam tarifi) kendi KUCUK st-key-gunkutusu_...
    kutusuna sahip (satir hizalamasi icin, bkz. _hafta_kartlarini_goster
    docstring'i). CSS de buna gore ayarlandi: SADECE SOL/SAG kenarlik
    (dikey sutun ayiricilari) -- boylece alt alta istiflenen kutular,
    ARADA BOSLUK OLMADAN, KESINTISIZ bir dikey cizgi gibi gorunuyor;
    UST/ALT kenarlik/radius/margin YOK (yatay ic-cizgi karmasasini
    onlemek icin). BILEREK buyuk paylasilan css_govdesi blogundan
    AYRI, KUCUK, IZOLE bir st.markdown cagrisi olarak tutuluyor --
    daha once bu CSS o buyuk bloga eklendiginde bir kismi sayfada
    DUZ METIN olarak SIZMISTI, kesin mekanizma belirlenemedi ama
    izole tutmak en dusuk riskli cozum.

    YUZ YEDINCI DUZELTME (5 Eylul 2026): Bahri "duzen tam istedigim
    gibi oldu, biraz guzellestirelim" dedi -- tabloyu CEVRELEYEN ust
    cizgi, Ogle/Aksam etiket satirlarina hafif zemin rengi, tarif
    linklerine/gun adi butonuna HOVER vurgusu ve biraz daha nefes
    alan ic bosluklar eklendi -- YAPI (satir hizalama, bosluksuz
    sutunlar) DEGISMEDI, sadece gorsel cila.

    YUZ SEKIZINCI DUZELTME (5 Eylul 2026): (1) tarih fontu, altindaki
    gun adi fontuyla (Fraunces, 600, 13.5px) AYNI yapildi -- eskiden
    daha kucuk/farkli bir font kullaniliyordu. (2) hafta tablosu
    kapsayicisi (st-key-haftatablosu_...) ICINDEKI TUM stVerticalBlock
    (Streamlit'in sutun govdesi) elemanlarinin gap'i 0'a cekildi --
    boylece ust uste yigilan kutular ARASINDA Streamlit'in kendi
    varsayilan boslugu KALMIYOR, kenarliklar/hafta sonu rengi KESINTISIZ
    gorunuyor. Bu kural SADECE bu tabloya ozel kapsayici icinde
    calisiyor, sayfadaki BASKA st.columns() kullanimlarini etkilemiyor.

    YUZ DOKUZUNCU DUZELTME (5 Eylul 2026, DUZELTME): Bahri "hala kesik
    cizgi ve hafta sonu arasinda beyaz boslukar var" dedi. Gercek
    render edilmis DOM incelenerek kok sebep bulundu: bir onceki
    kuralim SADECE haftatablosu kapsayicisinin ICINDEKI (descendant)
    stVerticalBlock'lari hedefliyordu -- ama asil bosluk, kapsayicinin
    KENDI DOGRUDAN cocuklari (her SATIR icin ayri bir stHorizontalBlock,
    ör. tarih satiri / gun adi satiri / her tarif satiri) ARASINDAKI
    gap'ten geliyordu. Kapsayicinin KENDISINE de gap:0 uygulanarak
    duzeltildi."""
    st.markdown(
        """
        <style>
        div[class*="st-key-haftatablosu_"] {
            gap: 0 !important;
        }
        div[class*="st-key-haftatablosu_"] div[data-testid="stVerticalBlock"] {
            gap: 0 !important;
        }
        div[class*="st-key-gunkutusu_"] {
            border-left: 1px solid #E4DDCB; border-right: 1px solid #E4DDCB;
            padding: 3px 7px; margin: 0; transition: background 0.15s ease;
        }
        div[class*="st-key-gunkutusu_hs_"] { background: #FBF0DC; }
        div[class*="st-key-gunkutusu_"] button {
            background: transparent !important; border: none !important;
            border-bottom: 2px solid #C88A2E !important; border-radius: 0 !important;
            padding: 5px 2px !important; width: 100%; box-shadow: none !important;
            transition: background 0.15s ease;
        }
        div[class*="st-key-gunkutusu_"] button:hover {
            background: rgba(200,138,46,0.12) !important;
        }
        div[class*="st-key-gunkutusu_"] button p {
            font-family: 'Fraunces', serif !important; font-weight: 600 !important;
            font-size: 13.5px !important; color: #2B2320 !important;
        }
        .omgo-tablo-tarih-hucre {
            padding: 6px 2px 3px; text-align: center; border-top: 2px solid #C88A2E;
            margin-top: -3px; font-family: 'Fraunces', serif; font-weight: 600;
            font-size: 13.5px; color: #2B2320;
        }
        .omgo-tablo-ogun-etiketi {
            font-size: 10px; font-weight: 700; text-transform: uppercase;
            letter-spacing: 0.05em; color: #C88A2E; text-align: center;
            background: rgba(200,138,46,0.08); padding: 3px 0; margin: 3px -7px 2px;
        }
        .omgo-tablo-bos-hucre { min-height: 20px; }
        div[class*="st-key-gunkutusu_"] div[data-testid="stPageLink"] {
            padding: 0; margin: 0; border-radius: 4px; transition: background 0.15s ease;
        }
        div[class*="st-key-gunkutusu_"] div[data-testid="stPageLink"]:hover {
            background: rgba(200,138,46,0.10);
        }
        div[class*="st-key-gunkutusu_"] div[data-testid="stPageLink"] p {
            font-size: 11px !important; padding: 1px 3px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _yillik_menu_tasarim_stilini_uygula():
    """YIRMINCI DUZELTME (13 Agustos 2026, Oturum 11): kullanicinin
    onayladigi kart-tasarim mockup'ini (tasarim_onizleme.html) GERCEK
    Streamlit widget'larina uyguluyoruz. Google Fonts (Fraunces baslik,
    Inter govde, IBM Plex Mono veri) + sicak kirectasi/patlican moru
    renk paleti -- tek seferde, sayfa basina enjekte edilir.

    NOT: Streamlit'in kendi buton/container DOM yapisina custom CSS
    uygulamak icin `key=` parametresiyle olusan `st-key-{key}` sinifina
    guveniyoruz (Streamlit >=1.38). CSS attribute-selector (*=) ile
    ORTAK bir on-ek (ör. "kart_") tasiyan TUM kartlari tek kuralla
    hedefliyoruz -- boylece her gun/hafta icin ayri CSS yazmaya gerek
    kalmiyor."""
    css_govdesi = """
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,400;9..144,600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
    <style>
    [data-testid='stPageLink'] p { white-space: normal !important; word-break: break-word !important; }
    div[class*="st-key-kart_"] { background: #EDE6D6; border-radius: 10px; box-shadow: 0 4px 12px rgba(43,35,32,0.10), 0 1px 2px rgba(43,35,32,0.08); padding: 0 0 4px; margin-bottom: 2px; overflow: hidden; }
    div[class*="st-key-kart_hs_"] { background: #F2CFA0; border: 2px solid #C88A2E; }
    div[class*="st-key-kartarka_"] { background: #3D2A3B; border-radius: 10px; box-shadow: 0 4px 12px rgba(43,35,32,0.18); padding: 0 0 4px; margin-bottom: 2px; overflow: hidden; }
    div[class*="st-key-kartarka_"] * { color: #EDE6D6 !important; }
    div[class*="st-key-baslik_"] button { width: 100%; background: transparent !important; border: none !important; border-bottom: 2px solid #C88A2E !important; border-radius: 0 !important; padding: 8px 10px 6px !important; text-align: left !important; box-shadow: none !important; white-space: pre-line !important; line-height: 1.2 !important; }
    div[class*="st-key-baslik_"] button p { font-family: 'Fraunces', serif !important; font-size: 15px !important; font-weight: 600 !important; color: #2B2320 !important; white-space: pre-line !important; line-height: 1.2 !important; }
    @media (max-width: 700px) { div[class*="st-key-baslik_"] button p { font-size: 10px !important; line-height: 1.15 !important; } div[class*="st-key-baslik_"] button { padding: 5px 4px 4px !important; } }
    @media (max-width: 480px) { div[class*="st-key-baslik_"] button p { font-size: 8.5px !important; } }
    /* YIRMI SEKIZINCI DUZELTME (13 Agustos 2026, Oturum 11): kullanici,
    mobil VE TABLET genislikte kartlarin (hem hafta gunu hem bolge menu
    kartlari) esit-sabit sutun sayisina sikisip metnin kelime ortasindan
    bolundugunu bildirdi. Onceki esik (700px) SADECE telefon genisligini
    kapsiyordu -- tablet (tipik 700-1024px) disarida kaliyordu, tam
    kullanicinin gordugu sorun buydu. Esik 1024px'e genisletildi, boylece
    tabletler de bu sutunlarin ALT SATIRA kaymasindan (flex-wrap)
    faydalaniyor -- 7 sutunluk hafta artik dar ekranlarda 3-4'erli
    satirlara bolunebiliyor, metin kelime ortasindan kesilmiyor. */
    @media (max-width: 1024px) { [data-testid="stHorizontalBlock"] { flex-wrap: wrap !important; row-gap: 10px; } [data-testid="stHorizontalBlock"] > div { flex: 1 1 30% !important; min-width: 130px !important; } }
    div[class*="st-key-cevir_"] button { background: #C88A2E !important; border: none !important; box-shadow: 0 2px 6px rgba(200,138,46,0.35) !important; color: #FFFFFF !important; font-size: 13px !important; font-weight: 600 !important; padding: 8px 16px !important; border-radius: 8px !important; }
    div[class*="st-key-cevir_"] button:hover { background: #B37B28 !important; }
    div[class*="st-key-popupkart_arka_"] div[class*="st-key-cevir_"] button { background: #E8B34A !important; color: #3D2A3B !important; box-shadow: 0 2px 6px rgba(0,0,0,0.25) !important; }
    div[class*="st-key-popupkart_arka_"] div[class*="st-key-cevir_"] button:hover { background: #F0C36A !important; }
    .omgo-veri-tablo { width: 100%; border-collapse: collapse; font-family: 'IBM Plex Mono', monospace; font-size: 12.5px; }
    .omgo-veri-tablo td { padding: 3px 0; }
    .omgo-veri-tablo td:last-child { text-align: right; font-weight: 500; }
    .omgo-veri-bolum { font-family: Inter, sans-serif; font-size: 11px; font-weight: 600; letter-spacing: 0.06em; text-transform: uppercase; color: #C88A2E; margin: 10px 0 3px; }
    .omgo-hedef-araligi { font-size: 10.5px; opacity: 0.65; font-style: italic; }
    /* DOKSANINCI DUZELTME (4 Eylul 2026): hedef disi kalan tablo
       satirlari icin -- Bahri'nin talebi: sadece rozette degil, ASIL
       TABLODA da hangi satir(lar) oldugu belirgin olsun.
       DOKSAN BIRINCI DUZELTME (4 Eylul 2026): ilk secilen acik pembe
       Bahri'ye uymadi -- uygulamanin ZATEN kullandigi "Hedef dışı"
       rozet rengiyle (rgba(166,71,47,...) kiremit/pas tonu) AYNI
       renk ailesine gecildi, boylece rozet ve tablo satiri tutarli
       (ayni "hedef disi" dili) gorunuyor -- koyu popup zemininde
       okunabilirlik icin acik krem renkli (#FDF6EC) metinle. */
    .omgo-satir-hedefdisi td { background: #A6472F !important; color: #FDF6EC !important; font-weight: 700; }
    .omgo-satir-hedefdisi .omgo-hedef-araligi { color: #FDF6EC !important; opacity: 0.85; }
    .omgo-ogun-baslik-buyuk { font-family: 'Fraunces', serif; font-size: 19px; font-weight: 700; text-align: center; color: #C88A2E; margin: 14px 0 8px; text-transform: uppercase; letter-spacing: 0.02em; }
    .omgo-hedef-rozet { display: inline-block; font-size: 11px; font-weight: 600; padding: 2px 9px; border-radius: 20px; margin-top: 6px; }
    .omgo-maliyet-baslik { font-family: 'Fraunces', serif; font-size: 15px; font-weight: 700; color: #7A531C; margin: 0 0 6px; }
    div[class*="st-key-maliyetkutu_"] { background: #EDE6D6; border: 1px solid #C88A2E; border-radius: 10px; padding: 12px 14px; margin: 10px 0; }
    .omgo-maliyet-tablo td { font-size: 14px !important; font-weight: 600 !important; color: #2B2320 !important; padding: 4px 0 !important; }
    div[class*="st-key-popupkart_arka_"] .omgo-maliyet-baslik { color: #7A531C !important; }
    div[class*="st-key-popupkart_arka_"] .omgo-maliyet-tablo td { color: #2B2320 !important; font-weight: 600 !important; font-size: 14px !important; }
    .omgo-veri-yok { font-size: 11.5px; font-style: italic; opacity: 0.65; margin-bottom: 6px; }
    .omgo-hedefte { background: rgba(91,117,83,0.30); color: #1B4D1B !important; }
    .omgo-hedefdisi { background: rgba(166,71,47,0.30); color: #6B2314 !important; }
    @keyframes omgoFlipOn { 0% { transform: rotateY(90deg); } 100% { transform: rotateY(0deg); } }
    @keyframes omgoFlipArka { 0% { transform: rotateY(-90deg); } 100% { transform: rotateY(0deg); } }
    [data-testid="stDialog"] { perspective: 1200px; }
    .stApp { perspective: 1200px; }
    [data-testid="stDialog"] h1, [data-testid="stDialog"] h2, [data-testid="stDialog"] [data-testid="stMarkdownContainer"] h1 { text-align: center !important; width: 100% !important; }
    div[class*="st-key-popupkart_on_"] { background: #EDE6D6; border-radius: 14px; padding: 18px 20px; box-shadow: 0 10px 30px rgba(43,35,32,0.20); animation: omgoFlipOn 0.65s cubic-bezier(0.3,0.1,0.2,1) both; transform-style: preserve-3d; backface-visibility: hidden; transform-origin: center center; }
    div[class*="st-key-popupkart_arka_"] { background: #3D2A3B; border-radius: 14px; padding: 18px 20px; box-shadow: 0 10px 30px rgba(43,35,32,0.30); animation: omgoFlipArka 0.65s cubic-bezier(0.3,0.1,0.2,1) both; transform-style: preserve-3d; backface-visibility: hidden; transform-origin: center center; }
    div[class*="st-key-popupkart_arka_"] * { color: #EDE6D6 !important; }
    div[class*="st-key-popupkart_arka_"] .omgo-veri-bolum { color: #C88A2E !important; }
    div[class*="st-key-popupkart_arka_"] .omgo-ogun-baslik-buyuk { color: #C88A2E !important; }
    </style>
    """
    # ONEMLI: Markdown, 4+ BOSLUKLA BASLAYAN HER SATIRI "kod blogu" sayip
    # oldugu gibi (kacis karakterleriyle) duz metin olarak gosterir --
    # unsafe_allow_html=True bile bunu gecersiz kilmaz, cunku bu karar HTML'e
    # ulasmadan ONCE, markdown asamasinda veriliyor. textwrap.dedent() SADECE
    # TUM satirlarin ORTAK asgari girintisini siler -- CSS kuralinin kendi
    # ic girintisi (ic ice suslu parantezler icin) hala 4+ bosluk birakabilir.
    # Bu yuzden tek satirlik CSS kurallari kullanildi VE guvenlik icin her
    # satir tek tek sifira indiriliyor (lstrip) -- boylece HICBIR satirda
    # 4+ bosluk kalma riski olmuyor. (13 Agustos 2026: canli ortamda CSS'in
    # duz metin olarak goruldugu hata bu sekilde duzeltildi.)
    css_temiz = "\n".join(satir.lstrip() for satir in css_govdesi.split("\n"))
    st.markdown(css_temiz, unsafe_allow_html=True)

def _gun_popup_govdesini_ciz(gun, detay, hedefler, fiyat_verisi_var, card_id, baslik_metni, hafta=None):
    """Pop-up icindeki TUM GORUNUR KARTI (baslik + on yuz YA DA arka
    yuz icerigi BIRLIKTE, TEK bir kapsayici icinde) cizer. YIRMI
    DORDUNCU DUZELTME (13 Agustos 2026): kullanici "sadece yazilar
    donuyor, kartin TAMAMI donmeli" dedi -- onceki versiyon sadece IC
    icerik bloguna animasyon uyguluyordu, baslik ve kartin kendi
    zemin/golge/koseleri (Streamlit'in dialog cercevesi disinda, BENIM
    cizdigim gorsel "kart" kismi) sabit kaliyordu. Simdi baslik + on/
    arka yuz icerigi TEK bir kapsayicida birlesip, animasyon bu
    kapsayicinin TAMAMINA uygulaniyor -- boylece gorunen "kart"in
    (Streamlit'in kendi dialog X-butonu/cercevesi disindaki HER SEY)
    butunu birlikte donuyor.

    Container key'i (popupkart_on_/popupkart_arka_ + card_id) on/arka
    yuz arasinda degistigi icin tarayici bunu HER SEFERINDE YENI bir
    eleman olarak gorur ve CSS @keyframes animasyonu HER cevirmede
    guvenilir sekilde yeniden calisir."""
    yuz_key = "yillik_menu_popup_yuz"
    st.session_state.setdefault(yuz_key, "on")

    # YIRMI ALTINCI DUZELTME (13 Agustos 2026, Oturum 11) -- kullanicinin
    # secimi "D": arka yuzde varsayilan olarak temel 5 deger + (varsa)
    # kullanicinin O AN hedefledigi ek besin ogeleri gosteriliyor;
    # ayrica "tum besin degerlerini gor" ile 32 ogenin tamami (Vitamin/
    # Mineral gruplarina ayrilmis) acilip kapatilabiliyor.
    # (TEMEL_5 artik modul seviyesinde tanimli, bkz. dosya basi.)
    VITAMIN_ANAHTARLARI = [
        "vitamin_a_mcg", "vitamin_b1_mg", "vitamin_b2_mg", "vitamin_b3_mg",
        "vitamin_b5_mg", "vitamin_b6_mg", "vitamin_b7_mcg", "vitamin_b9_mcg",
        "vitamin_b12_mcg", "vitamin_c_mg", "vitamin_d_mcg", "vitamin_e_mg", "vitamin_k_mcg",
    ]
    MINERAL_ANAHTARLARI = [
        "kalsiyum_mg", "demir_mg", "magnezyum_mg", "potasyum_mg", "cinko_mg",
        "fosfor_mg", "bakir_mg", "manganez_mg", "selenyum_mcg", "iyot_mcg",
    ]
    DIGER_MAKRO_ANAHTARLARI = ["sodyum_mg", "lif_g", "seker_g", "doymus_yag_g"]

    def _deger_formatla(deger):
        if deger is None:
            return "-"
        return f"{deger:.2f}" if deger < 5 else f"{round(deger)}"

    def _birim_al(anahtar):
        etiket = BESIN_ETIKET.get(anahtar, "")
        if "(" in etiket:
            return etiket[etiket.rfind("(") + 1: etiket.rfind(")")]
        return ""

    def _kisa_ad(anahtar):
        etiket = BESIN_ETIKET.get(anahtar, anahtar)
        return etiket.split(" (")[0].replace("Vitamin ", "")

    def _tablo_satirlari_yaz(anahtarlar, t, ogun_hedefleri=None, basarisiz_anahtarlar=None):
        """Satirlari yazar, en az bir GERCEK deger yazilip yazilmadigini
        (True/False) dondurur -- YIRMI YEDINCI DUZELTME (13 Agustos 2026):
        kullanici, bir kategoride HIC veri olmadiginda basligin altinin
        sessizce bos kalmasinin "bozuk" gibi gorundugunu bildirdi (halbuki
        gercek sebep: o gunun secilen yemeklerindeki malzemelerin bu
        belirli besin ogeleri icin katalogumuzda henuz veri OLMAMASI --
        SQL ile dogrulandi, ör. 564 malzemeden sadece 68'inde Vitamin C
        var). Artik veri yoksa acikca "Bu degerler icin veri yok" notu
        gosteriliyor, sessiz bosluk birakilmiyor.

        YETMIS SEKIZINCI DUZELTME (3 Eylul 2026): ogun_hedefleri verilirse
        (anahtar -> (alt, ust)), her satirin yanina -- varsa -- hedef
        araligi da yaziliyor, boylece deger ve hedef bir arada gorulebiliyor.

        DOKSANINCI DUZELTME (4 Eylul 2026): basarisiz_anahtarlar verilirse,
        o listede olan satirlar 'omgo-satir-hedefdisi' sinifiyla KIRMIZI
        vurgulanıyor -- Bahri'nin talebi: sadece rozette degil, ASIL
        TABLODA da hangi satirin hedef disi oldugu belirgin olsun."""
        ogun_hedefleri = ogun_hedefleri or {}
        basarisiz_anahtarlar = basarisiz_anahtarlar or []
        satirlar = []
        for anahtar in anahtarlar:
            deger = t.get(anahtar)
            if deger is None:
                continue
            aralik = ogun_hedefleri.get(anahtar)
            hedef_metni = f" <span class='omgo-hedef-araligi'>({aralik[0]}–{aralik[1]})</span>" if aralik else ""
            satir_sinifi = " class='omgo-satir-hedefdisi'" if anahtar in basarisiz_anahtarlar else ""
            satirlar.append(
                f"<tr{satir_sinifi}><td>{_kisa_ad(anahtar)}</td><td>{_deger_formatla(deger)} {_birim_al(anahtar)}{hedef_metni}</td></tr>"
            )
        if satirlar:
            st.markdown("<table class='omgo-veri-tablo'>" + "".join(satirlar) + "</table>", unsafe_allow_html=True)
            return True
        st.markdown(
            "<div class='omgo-veri-yok'>Bu değerler için veri yok (seçilen yemeklerin "
            "malzemelerinde henüz ölçülmemiş)</div>",
            unsafe_allow_html=True,
        )
        return False

    kapsayici_key = f"popupkart_{'on' if st.session_state[yuz_key]=='on' else 'arka'}_{card_id}"
    with st.container(key=kapsayici_key):
        st.markdown(
            f"<div style='font-family:Fraunces,serif; font-size:22px; font-weight:600; "
            f"color:{'#2B2320' if st.session_state[yuz_key]=='on' else '#EDE6D6'}; "
            f"white-space:pre-line; margin-bottom:8px;'>{baslik_metni}</div>",
            unsafe_allow_html=True,
        )

        if st.session_state[yuz_key] == "on":
            for ogun_adi, tarif_adlari in gun["ogunler"].items():
                st.markdown(
                    f"<div class='omgo-ogun-baslik-buyuk'>{ogun_adi.upper()} YEMEĞİ</div>",
                    unsafe_allow_html=True,
                )
                for ad in tarif_adlari:
                    st.page_link(
                        "pages/5_Tarif_Kutuphanesi.py", label=ad,
                        query_params={"tarif": ad}, use_container_width=True,
                    )
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
            with st.container(key=f"cevir_{card_id}"):
                if st.button("◤ Besin değerleri ve maliyet", key=f"btn_cevir_{card_id}", use_container_width=True):
                    st.session_state[yuz_key] = "arka"
                    st.rerun()
        else:
            # O an hedeflenen (kullanicinin "Ogun basina besin hedefi
            # uygula" ile sectigi) ek anahtarlar -- temel 5'in disinda
            # kalanlar. Hangi ogun icin oldugu farketmeksizin, HER iki
            # ogunde de ayni ek anahtarlar gosterilir (hedefler genelde
            # her iki ogun icin de ayni besin ogelerini kapsar).
            hedeflenen_ek_anahtarlar = []
            if hedefler:
                for ogun_hedefleri in hedefler.values():
                    for anahtar in ogun_hedefleri:
                        if anahtar not in TEMEL_5 and anahtar not in hedeflenen_ek_anahtarlar:
                            hedeflenen_ek_anahtarlar.append(anahtar)

            # YIRMI DOKUZUNCU/OTUZ DORDUNCU DUZELTME (13 Agustos 2026,
            # Oturum 11): kullanicinin son netlestirmesiyle -- BESIN
            # DEGERLERI musterinin gercekte yedigi 1 PORSIYON uzerinden
            # gosterilmeli (saglik/beslenme bilgisi olarak dogru
            # olan budur), ama MALIYET mutfak/uretim planlamasi icin
            # STANDART 10 PORSIYON uzerinden hesaplanmaya devam ediyor.
            # Bu yuzden SADECE maliyet_eur olcekleniyor, kalori/protein/
            # yag/karbonhidrat ve tum 27 genisletilmis besin ogesi
            # OLCEKLENMEDEN (t_ham, 1 porsiyon) gosteriliyor -- ayrica
            # bu, hedeflenen_ek_anahtarlar tablosunu da otomatik olarak
            # dogru hale getiriyor, cunku hedef araliklari da 1 porsiyon
            # baz alinarak kalibre edilmisti (bkz. OTUZ IKINCI DUZELTME).
            #
            # SEKSEN BIRINCI DUZELTME (3 Eylul 2026): porsiyon profili
            # secimi pop-up'tan sayfa seviyesine (Yıl/Ay'ın yanina)
            # tasindi -- boylece HER pop-up'ta ayri ayri secmek yerine
            # bir kez secilip TUM gunler icin gecerli oluyor. Secim,
            # sayfanin en ustunde (bkz. asagida "sag2" sutunu) yapiliyor
            # ve st.session_state uzerinden buraya ulasiyor.
            PORSIYON_STANDART = st.session_state.get("secili_porsiyon_sayisi", 10)
            OLCEKLENECEK_ALANLAR = ["maliyet_eur"]

            for ogun_adi, tarif_adlari in gun["ogunler"].items():
                t_ham = _ogun_toplami(tarif_adlari, detay)
                t = dict(t_ham)
                for alan in OLCEKLENECEK_ALANLAR:
                    if t.get(alan) is not None:
                        t[alan] = t[alan] * PORSIYON_STANDART
                gi_metin = f"{t['gi']:.1f}" if t["gi"] is not None else "-"
                st.markdown(f"<div class='omgo-ogun-baslik-buyuk'>{ogun_adi.upper()} YEMEĞİ</div>", unsafe_allow_html=True)
                st.markdown(
                    "<div style='text-align:center; font-size:11px; opacity:0.7; "
                    "margin-top:-6px; margin-bottom:8px;'>(besin değerleri 1 porsiyon için)</div>",
                    unsafe_allow_html=True,
                )

                # DOKSANINCI DUZELTME (4 Eylul 2026): hedefte/basarisiz_anahtarlar
                # artik BURADA (tablo cizilmeden ONCE) hesaplaniyor -- Bahri'nin
                # talebi: sadece "Hedef dışı: X, Y" rozetinde degil, ASIL
                # TABLODA da hedef disi kalan satirlar (ör. Protein, Glisemik
                # İndeks) KIRMIZI ile belirgin gosterilsin. Ayni hesaplama
                # asagidaki rozette de TEKRAR KULLANILIYOR (iki kez hesaplamiyoruz).
                hedefte, basarisiz_anahtarlar = _hedefte_mi(ogun_adi, t_ham, hedefler, hafta, detay)

                def _satir_sinifi(anahtar):
                    return " class='omgo-satir-hedefdisi'" if anahtar in basarisiz_anahtarlar else ""

                # YETMIS SEKIZINCI DUZELTME (3 Eylul 2026): kullanicinin
                # istegiyle -- "Hedeflenen Degerler" basligi HER SEYIN
                # USTUNE alindi, VE artik SADECE ek besin ogelerini degil
                # TEMEL 5'i de (kalori/protein/yag/karbonhidrat/GI) icine
                # aliyor -- bunlar icin de ogun bazli hedef araligi
                # tanimlanabiliyor, o yuzden ayni "hedeflenen" mantigina
                # dahil edilmeleri dogru. Her satirda deger + (varsa)
                # hedef araligi yan yana gosteriliyor.
                _ogun_hedefleri = (hedefler or {}).get(ogun_adi) or {}

                def _hedef_metni(anahtar):
                    aralik = _ogun_hedefleri.get(anahtar)
                    return f" <span class='omgo-hedef-araligi'>({aralik[0]}–{aralik[1]})</span>" if aralik else ""

                st.markdown("<div class='omgo-veri-bolum'>HEDEFLENEN BESİN DEĞERLERİ</div>", unsafe_allow_html=True)
                # DOKSAN DORDUNCU DUZELTME (4 Eylul 2026): Bahri, "Yağ: 40 g,
                # hedef 10.0-40.0" GORUNDUGU HALDE "hedef dışı" isaretlendigini
                # gordu -- "imkansiz matematik hatasi" sandi. Gercek sebep:
                # round() TAM SAYIYA yuvarliyordu (ör. gercek deger 40.4 iken
                # "40" gosteriliyordu), ama kontrol ROUNDLANMAMIS gercek
                # degeri kullaniyor -- 40.4 > 40.0 oldugu icin DOGRU sekilde
                # hedef disi sayiliyordu, sadece GORUNTUDE bu fark
                # SAKLANIYORDU. Artik TUM TEMEL_5 degerleri 1 ondalik
                # hassasiyetle gosteriliyor -- kontrolle GORUNEN deger
                # birebir tutarli, boyle bir "gorsel paradoks" bir daha
                # olusmuyor.
                st.markdown(
                    "<table class='omgo-veri-tablo'>"
                    f"<tr{_satir_sinifi('kalori')}><td>Kalori</td><td>{t['kalori']:.1f} kcal{_hedef_metni('kalori')}</td></tr>"
                    f"<tr{_satir_sinifi('protein')}><td>Protein</td><td>{t['protein']:.1f} g{_hedef_metni('protein')}</td></tr>"
                    f"<tr{_satir_sinifi('yag')}><td>Yağ</td><td>{t['yag']:.1f} g{_hedef_metni('yag')}</td></tr>"
                    f"<tr{_satir_sinifi('karbonhidrat')}><td>Karbonhidrat</td><td>{t['karbonhidrat']:.1f} g{_hedef_metni('karbonhidrat')}</td></tr>"
                    f"<tr{_satir_sinifi('gi')}><td>Glisemik İndeks</td><td>{gi_metin}{_hedef_metni('gi')}</td></tr>"
                    "</table>",
                    unsafe_allow_html=True,
                )
                if hedeflenen_ek_anahtarlar:
                    _tablo_satirlari_yaz(hedeflenen_ek_anahtarlar, t, _ogun_hedefleri, basarisiz_anahtarlar)

                # OTUZ BESINCI DUZELTME (13 Agustos 2026, Oturum 11):
                # kullanicinin istegiyle -- Alerjen maliyetle ilgili
                # olmadigi icin besin verilerinin HEMEN ALTINA (Iyot'un
                # altina) tasindi; "Hedef dısı/Hedefte" rozeti de besin
                # degerlerinin hedefe uyup uymadigini gosterdigi icin
                # (maliyetle degil) Alerjen'in hemen altina alindi.
                # Maliyet bolumu artik bunlarin ALTINDA, ayri ve
                # belirgin bir baslikla.
                alerjen_metin = ", ".join(sorted(t["alerjenler"])) if t["alerjenler"] else "Yok"
                st.markdown(
                    f"<table class='omgo-veri-tablo'><tr><td>Alerjen</td><td>{alerjen_metin}</td></tr></table>",
                    unsafe_allow_html=True,
                )
                # ONEMLI: hedefler (kullanicinin "Ogun basina besin hedefi"
                # ile belirledigi araliklar) HEP 1 PORSIYON baz alinarak
                # tasarlandi -- bu yuzden hedef kontrolu OLCEKLENMEMIS
                # t_ham ile yapiliyor, ekranda gosterilen 10-porsiyonluk
                # t degil.
                #
                # SEKSEN ALTINCI DUZELTME (4 Eylul 2026): "Hedef dışı"
                # rozeti GERI GETIRILDI. SEKSENINCI DUZELTME'de (3 Eylul)
                # Bahri'nin "bir daha gormek istemiyorum" talebiyle
                # kaldirilmisti, ama bu kez Bahri BASKA bir sorunu
                # (hedef canli/yeniden hesaplaniyor, uretim aninda
                # kullanilan hedeften FARKLI olabiliyor -- bkz. SEKSEN
                # BESINCI DUZELTME notlari) rozet SESSIZCE gizlendigi
                # icin GEC fark etti ve rozetin GERI GELMESINI istedi.
                #
                # SEKSEN DOKUZUNCU DUZELTME (4 Eylul 2026): Bahri'nin
                # talebi -- "Hedef dışı" yazisinin yaninda HANGI besin
                # ogesinin (ör. B12) hedef disi oldugu da gosteriliyor.
                # Ayrica TEMEL_5 disindaki ogeler artik HAFTALIK
                # ORTALAMA uzerinden kontrol ediliyor (bkz. _hedefte_mi).
                # (hedefte/basarisiz_anahtarlar YUKARIDA, tablo cizilmeden
                # once zaten hesaplandi -- burada TEKRAR hesaplamiyoruz.)
                if hedefte is True:
                    st.markdown("<span class='omgo-hedef-rozet omgo-hedefte'>Hedefte</span>", unsafe_allow_html=True)
                elif hedefte is False:
                    _basarisiz_etiketler = ", ".join(BESIN_ETIKET.get(a, a) for a in basarisiz_anahtarlar)
                    st.markdown(
                        f"<span class='omgo-hedef-rozet omgo-hedefdisi'>Hedef dışı: {_basarisiz_etiketler}</span>",
                        unsafe_allow_html=True,
                    )

                with st.container(key=f"maliyetkutu_{card_id}_{ogun_adi}"):
                    st.markdown(
                        f"<div class='omgo-maliyet-baslik'>Maliyet ({PORSIYON_STANDART} porsiyon için)</div>",
                        unsafe_allow_html=True,
                    )
                    if not fiyat_verisi_var:
                        st.markdown(
                            "<table class='omgo-veri-tablo omgo-maliyet-tablo'>"
                            "<tr><td>Maliyet</td><td>-</td></tr>"
                            "</table>",
                            unsafe_allow_html=True,
                        )
                    elif not t["tam_fiyatli"]:
                        eksik_liste = ", ".join(sorted(t["eksik_malzemeler"]))
                        st.markdown(
                            "<table class='omgo-veri-tablo omgo-maliyet-tablo'>"
                            f"<tr><td>Maliyet</td><td>≈{t['maliyet_eur']:.2f} € (eksik: {eksik_liste})</td></tr>"
                            "</table>",
                            unsafe_allow_html=True,
                        )
                    else:
                        # OTUZ BIRINCI DUZELTME (13 Agustos 2026): Tarif
                        # Kutuphanesi'ndeki "Gerçek üretim maliyeti (malzeme +
                        # enerji + işçilik)" modeliyle TUTARLI olmasi icin,
                        # her yemegin uretim asamalarindan enerji+iscilik
                        # maliyeti de hesaplanip malzeme maliyetine ekleniyor.
                        tarif_id_sozluk = _tum_tarif_id_by_ad_getir(st.session_state.isletme_id)
                        ayarlar = _maliyet_ayarlarini_getir(st.session_state.isletme_id)
                        toplam_enerji = 0.0
                        toplam_iscilik = 0.0
                        for ad in tarif_adlari:
                            rid = tarif_id_sozluk.get(ad)
                            if rid is None:
                                continue
                            asamalar = _uretim_asamalarini_getir(rid)
                            if asamalar:
                                e, i = _gercek_maliyet_hesapla(asamalar, ayarlar, PORSIYON_STANDART)
                                toplam_enerji += e
                                toplam_iscilik += i
                        malzeme_eur = t["maliyet_eur"]  # zaten PORSIYON_STANDART ile olceklendi
                        toplam_eur = malzeme_eur + toplam_enerji + toplam_iscilik
                        st.markdown(
                            "<table class='omgo-veri-tablo omgo-maliyet-tablo'>"
                            f"<tr><td>Malzeme</td><td>{malzeme_eur:.2f} €</td></tr>"
                            f"<tr><td>Enerji</td><td>{toplam_enerji:.2f} €</td></tr>"
                            f"<tr><td>İşçilik</td><td>{toplam_iscilik:.2f} €</td></tr>"
                            f"<tr><td>Toplam Maliyet</td><td>{toplam_eur:.2f} €</td></tr>"
                            "</table>",
                            unsafe_allow_html=True,
                        )
                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
            with st.container(key=f"cevir_{card_id}"):
                # DOKSAN IKINCI DUZELTME (4 Eylul 2026): Bahri'nin talebi --
                # eski tek "Gunluk menuye don" dugmesi, "Tekrar Dene" +
                # "Devam Et" ikilisine cevrildi.
                #
                # DOKSAN BESINCI DUZELTME (4 Eylul 2026): Bahri iki sorun
                # bildirdi -- (1) "Tekrar Dene" HER IKI ogunu (Ogle+Aksam)
                # birden yeniden uretiyordu, bu yuzden ZATEN hedefte olan
                # ogun BOZULABILIYORDU; (2) defalarca manuel tiklamaya
                # ragmen duzelmiyordu -- (1)'in DOGRUDAN sonucu: her
                # tiklamada iyi olan da yeniden karistigi icin "duzelt-
                # bozdur" dongusune giriyordu. Cozum: artik SADECE
                # GERCEKTEN hedef disi olan ogun(ler) yeniden uretiliyor,
                # hedefteki ogun(ler)e HIC DOKUNULMUYOR. Ayrica tek ogun
                # basina otomatik deneme sayisi 15'ten 30'a cikarildi
                # (artik ayni anda IKI degil TEK ogun uretildigi icin
                # ayni sure butcesiyle daha fazla deneme sigar).
                _tekrar_dene_col, _devam_et_col = st.columns(2)
                with _tekrar_dene_col:
                    if st.button("🔄 Tekrar Dene", key=f"btn_tekrar_dene_{card_id}", use_container_width=True):
                        _yeniden_uretilecek_ogunler = []
                        for _ogun_adi_kontrol, _tarif_adlari_kontrol in (gun.get("ogunler") or {}).items():
                            _t_ham_kontrol = _ogun_toplami(_tarif_adlari_kontrol, detay)
                            _hedefte_kontrol, _ = _hedefte_mi(_ogun_adi_kontrol, _t_ham_kontrol, hedefler, hafta, detay)
                            if _hedefte_kontrol is False:
                                _yeniden_uretilecek_ogunler.append(_ogun_adi_kontrol)

                        if not _yeniden_uretilecek_ogunler:
                            st.info("Bu günde hedef dışı bir öğün yok -- yeniden üretilecek bir şey bulunamadı.")
                        else:
                            with st.spinner("Sorunlu öğün(ler) yeniden deneniyor..."):
                                _grup1 = [t for t in tarifler_zengin if t["grup"] == 1]
                                _grup2 = [t for t in tarifler_zengin if t["grup"] == 2]
                                _grup3 = [t for t in tarifler_zengin if t["grup"] == 3]
                                _grup4 = [t for t in tarifler_zengin if t["grup"] == 4]
                                _gun_mevsimi = _tarih_mevsimi(gun["tarih"]) if gun.get("tarih") else None

                                # DEGISMEYECEK (hedefte olan) ogunlerin +
                                # haftanin DIGER gunlerinin kullandigi TUM
                                # tarifleri "kullanilmis" say -- yeniden
                                # uretilen ogun bunlarla CAKISMASIN.
                                _kullanilan_hafta_disarida = set()
                                for _gun2 in (hafta or []):
                                    for _ogun_adi2, _tarif_adlari2 in (_gun2.get("ogunler") or {}).items():
                                        if _gun2 is gun and _ogun_adi2 in _yeniden_uretilecek_ogunler:
                                            continue
                                        for _ad in (_tarif_adlari2 or []):
                                            _kullanilan_hafta_disarida.add(_ad)

                                _kullanilan_gun_taban_disarida = set()
                                for _ogun_adi2, _tarif_adlari2 in (gun.get("ogunler") or {}).items():
                                    if _ogun_adi2 in _yeniden_uretilecek_ogunler:
                                        continue
                                    for _ad in (_tarif_adlari2 or []):
                                        _kullanilan_gun_taban_disarida.add(_taban_kelime(_ad))

                                for _ogun_adi_hedef in _yeniden_uretilecek_ogunler:
                                    _hedef2 = (hedefler or {}).get(_ogun_adi_hedef)
                                    _en_iyi_uclu = None
                                    _en_iyi_mesafe = None
                                    for _deneme in range(30):
                                        _rastgele_deneme = random.Random()
                                        _kullanilan_hafta_deneme = set(_kullanilan_hafta_disarida)
                                        _kullanilan_gun_taban_deneme = set(_kullanilan_gun_taban_disarida)
                                        _t1n, _t2n, _t3n = ogun_olustur(
                                            _grup1, _grup2, _grup3, _gun_mevsimi,
                                            _kullanilan_hafta_deneme, _rastgele_deneme,
                                            _hedef2, _kullanilan_gun_taban_deneme,
                                        )
                                        _mesafe = _hedef_mesafesi(_t1n, _t2n, _t3n, _hedef2)
                                        if _en_iyi_mesafe is None or _mesafe < _en_iyi_mesafe:
                                            _en_iyi_mesafe = _mesafe
                                            _en_iyi_uclu = (_t1n, _t2n, _t3n)
                                        if _en_iyi_mesafe == 0:
                                            break

                                    _t1b, _t2b, _t3b = _en_iyi_uclu
                                    _yeni_tarif_adlari = [_t1b["ad"], _t2b["ad"], _t3b["ad"]]
                                    if _grup4:
                                        _birlesik = set(_t1b["etiketler"]) | set(_t2b["etiketler"]) | set(_t3b["etiketler"])
                                        _t4b = _fast_food_sec(_grup4, _birlesik, random.Random())
                                        if _t4b is not None:
                                            _yeni_tarif_adlari.append(_t4b["ad"])

                                    gun["ogunler"][_ogun_adi_hedef] = _yeni_tarif_adlari
                                    # Ayni gun icinde BASKA bir ogun de
                                    # yeniden uretilecekse, bu YENI secilenler
                                    # de "kullanilmis" sayilmali (ör. hem
                                    # Ogle hem Aksam hedef disiysa, ikisi
                                    # ayni tarifi SECMESIN).
                                    for _ad in _yeni_tarif_adlari:
                                        _kullanilan_hafta_disarida.add(_ad)
                                        _kullanilan_gun_taban_disarida.add(_taban_kelime(_ad))
                            st.rerun()
                with _devam_et_col:
                    if st.button("✓ Devam Et", key=f"btn_cevir2_{card_id}", use_container_width=True):
                        st.session_state[yuz_key] = "on"
                        st.rerun()


@st.dialog("Günün Menüsü")
def _gun_popup_dialog(gun, detay, hedefler, fiyat_verisi_var, card_id, baslik_metni, hafta=None):
    _gun_popup_govdesini_ciz(gun, detay, hedefler, fiyat_verisi_var, card_id, baslik_metni, hafta)


def _hafta_kartlarini_goster(hafta, detay, fiyat_verisi_var, hedefler, ay_adi, hafta_no, yil_secimi=None):
    """Haftayi GERCEK bir tabloda gosterir -- YUZ ALTINCI DUZELTME
    (5 Eylul 2026): Bahri'nin uc net duzeltmesi:
    1) Tarih GUN ADININ USTUNE alinir (en ustte "29 Aralık", altinda
       "Pazartesi").
    2) Gunler arasindaki bosluklar kaldirilir -- st.columns(gap=None)
       + kutu CSS'inde kenarlik/dolgu sikilastirildi, gercek bir tablo
       gibi BITISIK gorunuyor.
    3) Tum hafta boyunca Öğle porsiyonlari YAN YANA HIZALI olmali,
       Akşam porsiyonlari da AYRI olarak yan yana hizali olmali.

    (3) icin ONEMLI bir mimari degisiklik gerekti: eskiden HER GUN
    kendi sutununda BAGIMSIZ olarak Ogle+Aksam listelerini alt alta
    diziyordu -- bir gunun 3, digerinin 2 tarifi varsa, Aksam bolumu
    gunler arasinda FARKLI yukseklikte baslıyordu (hizasiz gorunuyordu).
    Simdi SATIR BAZLI render ediliyor: HER "satir" (tarih satiri, gun
    adi satiri, Öğle 1. tarif satiri, Öğle 2. tarif satiri, ...) kendi
    st.columns(7) cagrisini alir -- boylece Streamlit'in kendi yatay
    flex duzeni sayesinde TUM gunlerin AYNI satirdaki icerigi GERCEKTEN
    ayni yukseklikte hizalanir. Eksik tarifi olan gunler icin o
    hucrede bos birakiliyor (hizayi bozmadan).

    YUZ SEKIZINCI DUZELTME (5 Eylul 2026): Bahri'nin gozlemi -- sutun
    aralarindaki/hafta sonu renklendirmesindeki cizgiler "kesik kesik"
    gorunuyordu. Kok sebep: Streamlit'in kendi st.columns() sutun
    govdesi (stVerticalBlock), icindeki ust uste yigilan kutularim
    arasina VARSAYILAN bir bosluk (gap) koyuyordu -- benim kendi
    kutularimin kenarliklari/arka plani DOGRU olsa bile, ARADAKI bu
    kucuk Streamlit-bosluğu yuzunden cizgiler/renk SUREKSIZ gorunuyordu.
    Cozum: TUM satirlari (tarih -> son aksam tarifi) SARAN bir dis
    kapsayici (st.container(key=...)) eklendi, CSS'te SADECE bu
    kapsayicinin ICINDEKI stVerticalBlock'larin gap'i 0'a cekildi --
    boylece degisiklik SADECE bu tabloyu etkiliyor, sayfadaki BASKA
    st.columns() kullanimlarina (ör. Abonelik sayfasi) DOKUNMUYOR."""
    _yillik_menu_tasarim_stilini_uygula()
    _tablo_stilini_uygula()

    GUN_ADLARI = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    card_idler = [f"{ay_adi}-{hafta_no}-{gun['gun']}" for gun in hafta]

    gun_bilgileri = []
    for gun in hafta:
        tarih = gun.get("tarih")
        if tarih is not None:
            # YUZ ONUNCU DUZELTME (5 Eylul 2026): Bahri Aralık->Ocak
            # (YIL sinirini asan) bir hafta gorunce, "1 Ocak" gibi
            # gorunen bir tarihin hangi YILA ait oldugunu anlayamadi --
            # yil hic gosterilmiyordu. Artik secili yildan (yil_secimi)
            # FARKLI bir yil ise, yil da ekleniyor (ör. "1 Ocak 2027").
            if yil_secimi is not None and tarih.year != yil_secimi:
                tarih_metni = f"{tarih.day} {AYLAR_SIRALI[tarih.month - 1]} {tarih.year}"
            else:
                tarih_metni = f"{tarih.day} {AYLAR_SIRALI[tarih.month - 1]}"
            gun_bilgileri.append({
                "gun_adi": GUN_ADLARI[tarih.weekday()],
                "hafta_sonu_mu": tarih.weekday() >= 5,
                "tarih_metni": tarih_metni,
            })
        else:
            gun_bilgileri.append({"gun_adi": f"Gün {gun['gun']}", "hafta_sonu_mu": False, "tarih_metni": ""})

    def _kutu_key(i, satir_no):
        hs = "hs_" if gun_bilgileri[i]["hafta_sonu_mu"] else ""
        return f"gunkutusu_{hs}{card_idler[i]}_{satir_no}"

    with st.container(key=f"haftatablosu_{ay_adi}_{hafta_no}"):
        # SATIR 1: Tarih (gun adindan ONCE -- Bahri'nin talebi)
        kolonlar = st.columns(len(hafta), gap=None)
        for i, kolon in enumerate(kolonlar):
            with kolon:
                with st.container(key=_kutu_key(i, "tarih")):
                    st.markdown(f"<div class='omgo-tablo-tarih-hucre'>{gun_bilgileri[i]['tarih_metni']}</div>", unsafe_allow_html=True)

        # SATIR 2: Gun adi (tiklaninca pop-up acan buton)
        kolonlar = st.columns(len(hafta), gap=None)
        for i, kolon in enumerate(kolonlar):
            with kolon:
                with st.container(key=_kutu_key(i, "gunadi")):
                    if st.button(gun_bilgileri[i]["gun_adi"], key=f"btn_gun_{card_idler[i]}", use_container_width=True):
                        st.session_state["yillik_menu_popup_gun_id"] = card_idler[i]
                        st.session_state["yillik_menu_popup_yuz"] = "arka"
                        st.rerun()

        # SATIR 3: "Öğle" etiketi
        kolonlar = st.columns(len(hafta), gap=None)
        for i, kolon in enumerate(kolonlar):
            with kolon:
                with st.container(key=_kutu_key(i, "ogle_etiket")):
                    st.markdown("<div class='omgo-tablo-ogun-etiketi'>Öğle</div>", unsafe_allow_html=True)

        # Öğle tarifleri -- SATIR BAZLI (max sayi kadar satir, her satirda TUM gunler)
        max_ogle = max((len(gun["ogunler"].get("Öğle", [])) for gun in hafta), default=0)
        for j in range(max_ogle):
            kolonlar = st.columns(len(hafta), gap=None)
            for i, (kolon, gun) in enumerate(zip(kolonlar, hafta)):
                with kolon:
                    with st.container(key=_kutu_key(i, f"ogle_{j}")):
                        liste = gun["ogunler"].get("Öğle", [])
                        if j < len(liste):
                            st.page_link(
                                "pages/5_Tarif_Kutuphanesi.py", label=liste[j],
                                query_params={"tarif": liste[j]}, use_container_width=True,
                            )
                        else:
                            st.markdown("<div class='omgo-tablo-bos-hucre'>&nbsp;</div>", unsafe_allow_html=True)

        # SATIR: "Akşam" etiketi
        kolonlar = st.columns(len(hafta), gap=None)
        for i, kolon in enumerate(kolonlar):
            with kolon:
                with st.container(key=_kutu_key(i, "aksam_etiket")):
                    st.markdown("<div class='omgo-tablo-ogun-etiketi'>Akşam</div>", unsafe_allow_html=True)

        # Akşam tarifleri -- ayni sekilde satir bazli
        max_aksam = max((len(gun["ogunler"].get("Akşam", [])) for gun in hafta), default=0)
        for j in range(max_aksam):
            kolonlar = st.columns(len(hafta), gap=None)
            for i, (kolon, gun) in enumerate(zip(kolonlar, hafta)):
                with kolon:
                    with st.container(key=_kutu_key(i, f"aksam_{j}")):
                        liste = gun["ogunler"].get("Akşam", [])
                        if j < len(liste):
                            st.page_link(
                                "pages/5_Tarif_Kutuphanesi.py", label=liste[j],
                                query_params={"tarif": liste[j]}, use_container_width=True,
                            )
                        else:
                            st.markdown("<div class='omgo-tablo-bos-hucre'>&nbsp;</div>", unsafe_allow_html=True)

    for i, gun in enumerate(hafta):
        card_id = card_idler[i]
        if st.session_state.get("yillik_menu_popup_gun_id") == card_id:
            # YUZ DOKUZUNCU DUZELTME (5 Eylul 2026): Bahri, pop-up'i bir
            # kere acip kapattigi halde SUREKLI tekrar actigini bildirdi.
            # Kok sebep: bu bayrak HICBIR ZAMAN temizlenmiyordu -- bir
            # butonun aksine (ki st.button() SADECE tikleme aninda True
            # doner, sonraki her calistirmada otomatik False'a doner),
            # benim session_state bayragim SONSUZA KADAR true kaliyordu.
            # Bu yuzden sayfadaki HERHANGI bir etkilesim (farkli bir
            # gune tiklamak DAHIL, hatta alakasiz bir sey) sayfayi
            # yeniden calistirdiginda, bu kosul HALA dogruydu ve
            # pop-up'i TEKRAR aciyordu.
            #
            # Resmi Streamlit dokumantasyonundan dogrulandi: st.dialog
            # kendi ICINDEKI etkilesimleri (Tekrar Dene/Devam Et gibi)
            # FRAGMENT olarak ele alir -- SADECE dialog fonksiyonunun
            # kendisini yeniden calistirir, DIS script'i (bu dongu dahil)
            # DEGIL. Yani bayragi BURADA, HEMEN tuketilir tuketilmez
            # temizlemek, ZATEN ACIK olan dialogun kendi ic
            # etkilesimlerini BOZMAZ -- sadece DIS script'in onu
            # GEREKSIZ YERE TEKRAR TEKRAR acmasini engeller.
            st.session_state["yillik_menu_popup_gun_id"] = None
            bilgi = gun_bilgileri[i]
            baslik_metni = f"{bilgi['gun_adi']}\n{bilgi['tarih_metni']}" if bilgi["tarih_metni"] else bilgi["gun_adi"]
            _gun_popup_dialog(gun, detay, hedefler, fiyat_verisi_var, card_id, baslik_metni, hafta)




def _aylik_menu_excel_olustur(aylik, detay, fiyat_verisi_var, hedefler):
    """Aylık menüyü ekrandaki kart görünümüyle AYNI düzende Excel'e döker:
    her gün bir sütun, altında Öğle/Akşam blokları (yemekler + besin +
    alerjen + maliyet) aynı sırayla. Bir finansal model degil -- formul
    gerekmiyor, sadece ekrandakiyle bire bir eslesen bir gorunum."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aylık Menü"

    yazi_tipi = "Arial"
    baslik_yazi = Font(name=yazi_tipi, bold=True, color="FFFFFF")
    baslik_dolgu = PatternFill(start_color="2C6B3C", end_color="2C6B3C", fill_type="solid")
    hafta_baslik_yazi = Font(name=yazi_tipi, bold=True, size=13)
    alan_yazi = Font(name=yazi_tipi, bold=True)
    normal_yazi = Font(name=yazi_tipi)
    RENK_ANA, RENK_YARDIMCI, RENK_TAMAMLAYICI, RENK_FAST_FOOD = "D85A30", "639922", "1D9E75", "BA7517"

    def oyun_bloguna_yaz(satir, ogun_adi, tarif_adlari, t, t_ham, gun_kolonu):
        ws.cell(row=satir, column=1, value=f"{ogun_adi} (besin: 1 p. / maliyet: 10 p.)").font = alan_yazi
        satir += 1
        # Yemek satirlari, o ogunde 4. (istege bagli Fast Food) tarif
        # var mi yok mu -- 6 Agustos 2026'da eklendi -- gore DINAMIK
        # olarak olusturuluyor. Sabit 3'lu bir liste kullanip
        # tarif_adlari[i] ile eslestirmek (eski kod), 4. tarif eklenince
        # onu sessizce Excel'den dusururdu.
        yemek_satirlari = [
            ("Ana Yemek", RENK_ANA), ("Yardımcı Yemek", RENK_YARDIMCI), ("Tamamlayıcı", RENK_TAMAMLAYICI),
        ]
        if len(tarif_adlari) >= 4:
            yemek_satirlari.append(("Fast Food", RENK_FAST_FOOD))
        yemek_sayisi = len(yemek_satirlari)

        alan_satirlari = yemek_satirlari + [
            ("Besin (kcal/P/Y/K/Gİ)", None), ("Alerjen", None), ("Maliyet", None),
        ]
        if hedefler:
            alan_satirlari.append(("Hedef Durumu", None))
        for i, (etiket, renk) in enumerate(alan_satirlari):
            hucre_etiket = ws.cell(row=satir, column=1, value=etiket)
            hucre_etiket.font = Font(name=yazi_tipi, color=renk) if renk else normal_yazi
            if i < yemek_sayisi:
                deger = tarif_adlari[i]
            elif etiket.startswith("Besin"):
                gi_deger = f"{round(t['gi'])}" if t["gi"] is not None else "-"
                deger = (
                    f"{round(t['kalori'])} kcal · P{round(t['protein'])}g · "
                    f"Y{round(t['yag'])}g · K{round(t['karbonhidrat'])}g · Gİ{gi_deger}"
                )
            elif etiket == "Alerjen":
                deger = ", ".join(sorted(t["alerjenler"])) if t["alerjenler"] else "Yok"
            elif etiket == "Hedef Durumu":
                hedefte, _ = _hedefte_mi(ogun_adi, t_ham, hedefler, hafta, detay)
                # SEKSEN ALTINCI DUZELTME (4 Eylul 2026): pop-up'taki
                # degisiklikle tutarli olsun diye "Hedef dışı" burada da
                # geri getirildi (bkz. yukaridaki not).
                deger = {True: "Hedefte", False: "Hedef dışı", None: "-"}[hedefte]
            else:  # Maliyet
                if not fiyat_verisi_var:
                    deger = "-"
                elif t["tam_fiyatli"]:
                    deger = f"{t['maliyet_eur']:.2f} €"
                else:
                    eksik_liste = ", ".join(sorted(t["eksik_malzemeler"]))
                    deger = f"≈{t['maliyet_eur']:.2f} € (eksik: {eksik_liste})"
            hucre = ws.cell(row=satir, column=gun_kolonu, value=deger)
            hucre.font = normal_yazi
            hucre.alignment = Alignment(wrap_text=True, vertical="top")
            satir += 1
        return satir

    satir = 1
    for hafta_no, hafta in enumerate(aylik["haftalar"], start=1):
        gun_sayisi = len(hafta)

        ws.cell(row=satir, column=1, value=f"{aylik['ay']} — {hafta_no}. Hafta").font = hafta_baslik_yazi
        satir += 1

        baslik_satiri = satir
        ws.cell(row=baslik_satiri, column=1, value="")
        for g in range(gun_sayisi):
            hucre = ws.cell(row=baslik_satiri, column=g + 2, value=f"Gün {g + 1}")
            hucre.font = baslik_yazi
            hucre.fill = baslik_dolgu
            hucre.alignment = Alignment(horizontal="center")
        satir += 1

        blok_baslangic = satir
        for g, gun in enumerate(hafta):
            gun_kolonu = g + 2
            s = blok_baslangic
            for ogun_adi, tarif_adlari in gun["ogunler"].items():
                # OTUZ DORDUNCU DUZELTME (13 Agustos 2026): pop-up'taki
                # ayni ilke -- besin degerleri 1 porsiyon (musterinin
                # gercekte yedigi), maliyet ise 10 porsiyon (mutfak
                # planlamasi) uzerinden. Hedefte kontrolu icin ayrica
                # olceklenmemis t_ham tutuluyor.
                t_ham = _ogun_toplami(tarif_adlari, detay)
                t = dict(t_ham)
                if t.get("maliyet_eur") is not None:
                    t["maliyet_eur"] = t["maliyet_eur"] * 10
                s = oyun_bloguna_yaz(s, ogun_adi, tarif_adlari, t, t_ham, gun_kolonu)
        satir = s + 1  # bir sonraki hafta bloğundan önce bos satir

    genislikler = [24] + [24] * 7
    for i, genislik in enumerate(genislikler, start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


aylik = st.session_state.get("yillik_menu_aylik")
if aylik:
    kayitli_hedefler = st.session_state.get("yillik_menu_hedefler")

    if aylik.get("gecmis_ay_mi"):
        st.warning(
            "Bu, geçmiş bir tarih için üretildi — gerçek servis kaydı "
            "değildir, sadece sistemin o dönem için ne önerdiğine dair "
            "bir örnektir."
        )

    # DOKSAN UCUNCU DUZELTME (4 Eylul 2026): "Aylık Menüyü Kaydet" --
    # Bahri'nin karari: SADECE TUM gunler hedefteyse (ya da hic hedef
    # tanimlanmamissa -- o zaman zaten ihlal edilecek bir sey yok)
    # aktif olsun, degilse uyari versin VE kaydetmeyi ENGELLESIN.
    #
    # YUZUNCU DUZELTME (4 Eylul 2026): tarihler artik HAM (tarih,
    # ogun_adi) ikilisi olarak toplanıyor -- eskiden burada dogrudan
    # ISO bicimli (2026-12-01) metin string'i olusturuluyordu, Bahri'nin
    # istedigi Turkce/gruplu bicimlendirme (bkz. _hedef_disi_liste_metni)
    # HAM veriye ihtiyac duyuyor.
    _hedef_disi_kayitlar = []
    if kayitli_hedefler:
        for _hafta in aylik["haftalar"]:
            for _gun in _hafta:
                for _ogun_adi, _tarif_adlari in (_gun.get("ogunler") or {}).items():
                    _t_ham = _ogun_toplami(_tarif_adlari, detay)
                    _hedefte_sonuc, _ = _hedefte_mi(_ogun_adi, _t_ham, kayitli_hedefler, _hafta, detay)
                    if _hedefte_sonuc is False:
                        _hedef_disi_kayitlar.append((_gun.get("tarih"), _ogun_adi))

    _secili_profil_id_kaydet = st.session_state.get("secili_porsiyon_profil_id")

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    if _hedef_disi_kayitlar:
        st.warning(
            f"Bu ayı kaydetmeden önce hedef dışı kalan {len(_hedef_disi_kayitlar)} "
            "öğünü \"Tekrar Dene\" butonu ile uygulamanın yaratacağı yeni öğünler "
            "ile gözden geçir, ve eğer limit aşımları tolerans dahilinde ise "
            "\"Devam Et\" butonu ile kabul et. Tüm hedef dışı günleri bitirince "
            "\"Aylık Menüyü Kaydet\" butonu ile bu ayın menüsünü kaydedersen, "
            "herhangi bir zamanda yeniden ulaşabilirsin.\n\n"
            + _hedef_disi_liste_metni(_hedef_disi_kayitlar[:16], yil_secimi, ay_secimi)
            + (" ..." if len(_hedef_disi_kayitlar) > 16 else "")
        )
        st.button("Aylık Menüyü Kaydet", disabled=True, key="btn_aylik_kaydet_disabled")
    elif not _secili_profil_id_kaydet:
        st.button("Aylık Menüyü Kaydet", disabled=True, key="btn_aylik_kaydet_disabled",
                   help="Önce yukarıdan bir porsiyon profili seç.")
    else:
        if st.button("💾 Aylık Menüyü Kaydet", key="btn_aylik_kaydet", type="primary"):
            _mevcut_kayit = (
                supabase.table("kayitli_aylik_menuler")
                .select("id")
                .eq("isletme_id", st.session_state.isletme_id)
                .eq("porsiyon_profil_id", _secili_profil_id_kaydet)
                .eq("yil", aylik["yil"])
                .eq("ay", aylik["ay"])
                .execute()
            ).data
            _kayit_govdesi = {
                "isletme_id": st.session_state.isletme_id,
                "porsiyon_profil_id": _secili_profil_id_kaydet,
                "yil": aylik["yil"],
                "ay": aylik["ay"],
                "menu_verisi": {"haftalar": aylik["haftalar"]},
            }
            if _mevcut_kayit:
                supabase.table("kayitli_aylik_menuler").update(_kayit_govdesi).eq("id", _mevcut_kayit[0]["id"]).execute()
                st.success(f"\"{aylik['ay']} {aylik['yil']}\" güncellenerek kaydedildi (önceki kayıt üzerine yazıldı).")
            else:
                supabase.table("kayitli_aylik_menuler").insert(_kayit_govdesi).execute()
                st.success(f"\"{aylik['ay']} {aylik['yil']}\" kaydedildi.")

    excel_verisi = _aylik_menu_excel_olustur(aylik, detay, fiyat_verisi_var, kayitli_hedefler)
    # OTUZ DORDUNCU DUZELTME (24 Agustos 2026): kod incelemesinde bulundu --
    # bu sayfada veritabanina yazma OLMADIGI icin "salt_okunur" hic
    # kullanilmamisti, ama bu buton bir ISTISNA: odeme onayi bekleyen
    # kullanici, sinirsiz sayida aylik menu uretip GERCEK, disari
    # tasinabilir bir Excel ciktisi alabiliyordu -- diger 3 sayfadaki
    # ("goruntule ama islem yapma") kurali fiilen boşa cikaran tek nokta
    # buydu. Menu ONIZLEMESI (ekrandaki kartlar) bilerek disabled
    # BIRAKILDI -- sadece disariya TASINABILIR/KALICI cikti (Excel)
    # engelleniyor, ayni Recete Uretimi/Ozel Menu Uretimi'ndeki "olustur/
    # kaydet" butonlarinin gated olup form alanlarinin gated olmamasi gibi.
    st.download_button(
        "Excel'e indir",
        data=excel_verisi,
        file_name=f"yillik_menu_{aylik['ay']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=st.session_state.get("salt_okunur", False),
    )

    fast_food_notasi = (
        "&nbsp;&nbsp;&nbsp;<span style='color:#BA7517;'>●</span> Fast Food"
        if st.session_state.get("kendi_menu_dahil") else ""
    )
    st.markdown(
        "<div style='font-size:13px; color:gray; margin:0.5rem 0 1rem;'>"
        "<span style='color:#D85A30;'>●</span> Ana Yemek&nbsp;&nbsp;&nbsp;"
        "<span style='color:#639922;'>●</span> Yardımcı Yemek&nbsp;&nbsp;&nbsp;"
        f"<span style='color:#1D9E75;'>●</span> Tamamlayıcılar{fast_food_notasi}</div>",
        unsafe_allow_html=True,
    )

    for i, hafta in enumerate(aylik["haftalar"], start=1):
        st.markdown(
            f"<div style='font-weight:700; margin:10px 0 4px; font-size:14px;'>"
            f"{aylik['ay']} — {i}. Hafta</div>",
            unsafe_allow_html=True,
        )
        _hafta_kartlarini_goster(hafta, detay, fiyat_verisi_var, kayitli_hedefler, aylik["ay"], i, aylik["yil"])
        # YUZUNCU DUZELTME (4 Eylul 2026): Bahri'nin "aralarda cok bosluk
        # var" gozlemi uzerine -- st.divider() (kalin, buyuk dikey bosluklu
        # bir cizgi) yerine COK INCE, DUSUK-MARJINLI ozel bir cizgi.
        st.markdown(
            "<hr style='margin:6px 0 10px; border:none; border-top:1px solid #DDD6C4;'>",
            unsafe_allow_html=True,
        )

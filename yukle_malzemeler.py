# yukle_malzemeler.py
#
# BİR KEREYE MAHSUS ETL scripti: kaynak_duzeltilmis_v2.xlsx dosyasındaki
# 337 malzemeyi + 14 alerjen ilişkisini Supabase'deki `malzemeler` ve
# `malzeme_alerjen` tablolarına yükler (varsayilan_fiyat_eur dahil).
#
# GÜVENLİK: Bu script SERVICE_ROLE anahtarını kullanır (RLS'i bypass eder).
# Bu anahtarı ASLA app.py/secrets.toml içine koyma, ASLA GitHub'a commitleme
# -- sadece bu scripti çalıştırırken, bir kereye mahsus gir.
#
# Kullanım (cmd):
#   set SUPABASE_URL=https://xxxx.supabase.co
#   set SUPABASE_SERVICE_ROLE_KEY=xxxx
#   python yukle_malzemeler.py kaynak_duzeltilmis_v2.xlsx
#
# (Ortam degiskenlerini onceden ayarlamazsan script calisirken senden ister.)

import os
import sys

import openpyxl
from supabase import create_client

SUPABASE_URL = os.environ.get("SUPABASE_URL") or input("SUPABASE_URL: ").strip()
SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or input(
    "SUPABASE_SERVICE_ROLE_KEY (Settings > API > Legacy anon, service_role API keys): "
).strip()
XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "kaynak_duzeltilmis_v2.xlsx"

supabase = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

ALLERGEN_COLS = [
    "GLUTEN", "KABUKLU DENİZ ÜRÜNÜ", "YUMURTA", "BALIK", "YER FISTIĞI", "SOYA",
    "SÜT", "SERT KABUKLU YEMİŞ", "KEREVİZ", "HARDAL", "SUSAM", "SÜLFİT (SO2)",
    "YUMUŞAKÇA", "LUPİN",
]
# xlsx sutun basligi -> alerjenler tablosundaki 'ad' degeri (schema'da
# Turkce karakterler sadelestirilerek seed edildigi icin bu esleme gerekli)
ALLERGEN_COL_TO_DB_AD = {
    "GLUTEN": "Gluten",
    "KABUKLU DENİZ ÜRÜNÜ": "Kabuklu Deniz Urunu",
    "YUMURTA": "Yumurta",
    "BALIK": "Balik",
    "YER FISTIĞI": "Yer Fistigi",
    "SOYA": "Soya",
    "SÜT": "Sut",
    "SERT KABUKLU YEMİŞ": "Sert Kabuklu Yemis",
    "KEREVİZ": "Kereviz",
    "HARDAL": "Hardal",
    "SUSAM": "Susam",
    "SÜLFİT (SO2)": "Sulfit (SO2)",
    "YUMUŞAKÇA": "Yumusakca",
    "LUPİN": "Lupin",
}

FIELD_COLS = [
    "yogunluk", "ozgul_isi", "bozulma", "fire", "saklama", "kalori", "protein",
    "yag", "kh", "gi", "mevsim", "fiyat", "isi_iletkenlik", "yuzey", "not",
]

CHUNK = 50


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["kaynak"]

    print("Dosya okunuyor...")
    rows = []
    current_cat_id = None

    for r in range(3, ws.max_row + 1):
        kategori_metni = ws.cell(row=r, column=1).value
        ad = ws.cell(row=r, column=2).value

        if kategori_metni:
            if str(kategori_metni).strip().upper() == "AÇIKLAMA":
                break  # veri bitti, sayfanin altindaki aciklama/legend basliyor
            current_cat_id = int(str(kategori_metni).split(".")[0].strip())

        if not ad or not str(ad).strip():
            continue

        degerler = {}
        for i, alan in enumerate(FIELD_COLS):
            degerler[alan] = ws.cell(row=r, column=3 + i).value

        allergen_start_col = 3 + len(FIELD_COLS)
        allerjenler_bu_satir = []
        for i, col_adi in enumerate(ALLERGEN_COLS):
            deger = ws.cell(row=r, column=allergen_start_col + i).value
            if deger == "X":
                allerjenler_bu_satir.append(ALLERGEN_COL_TO_DB_AD[col_adi])

        rows.append(
            {
                "kategori_id": current_cat_id,
                "ad": str(ad).strip(),
                "yogunluk": degerler["yogunluk"],
                "ozgul_isi": degerler["ozgul_isi"],
                "bozulma_suresi": degerler["bozulma"],
                "fire_orani": degerler["fire"],
                "saklama_isisi": degerler["saklama"],
                "kalori": degerler["kalori"],
                "protein": degerler["protein"],
                "yag": degerler["yag"],
                "karbonhidrat": degerler["kh"],
                "glisemik_indeks": degerler["gi"],
                "mevsim": degerler["mevsim"],
                "isi_iletkenlik": degerler["isi_iletkenlik"],
                "yuzey_alani": degerler["yuzey"],
                "not_aciklama": degerler["not"],
                "varsayilan_fiyat_eur": degerler["fiyat"],
                "_allerjenler": allerjenler_bu_satir,
            }
        )

    print(f"{len(rows)} malzeme okundu. Supabase'e yükleniyor...")

    eklenecekler = []
    for r in rows:
        satir = {k: v for k, v in r.items() if k != "_allerjenler"}
        satir["isletme_id"] = None
        eklenecekler.append(satir)

    for i in range(0, len(eklenecekler), CHUNK):
        parca = eklenecekler[i : i + CHUNK]
        supabase.table("malzemeler").insert(parca).execute()
        print(f"  {i + len(parca)}/{len(eklenecekler)} malzeme eklendi")

    print("Malzeme-alerjen ilişkileri kuruluyor...")

    alerjen_kayitlari = supabase.table("alerjenler").select("id, ad").execute().data
    alerjen_id_by_ad = {a["ad"]: a["id"] for a in alerjen_kayitlari}

    malzeme_kayitlari = (
        supabase.table("malzemeler").select("id, ad").is_("isletme_id", "null").execute().data
    )
    malzeme_id_by_ad = {m["ad"]: m["id"] for m in malzeme_kayitlari}

    iliski_eklenecekler = []
    for r in rows:
        malzeme_id = malzeme_id_by_ad.get(r["ad"])
        if not malzeme_id:
            continue
        for alerjen_ad in r["_allerjenler"]:
            alerjen_id = alerjen_id_by_ad.get(alerjen_ad)
            if alerjen_id:
                iliski_eklenecekler.append({"malzeme_id": malzeme_id, "alerjen_id": alerjen_id})

    for i in range(0, len(iliski_eklenecekler), CHUNK):
        parca = iliski_eklenecekler[i : i + CHUNK]
        supabase.table("malzeme_alerjen").insert(parca).execute()

    print(f"Tamamlandı: {len(eklenecekler)} malzeme, {len(iliski_eklenecekler)} alerjen ilişkisi eklendi.")


if __name__ == "__main__":
    main()
